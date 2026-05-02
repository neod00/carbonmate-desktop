"""
PoC — 산정 데이터 xlsx 자동 생성기 (4시트: 표지 / LCIA / 민감도 분석 / 사용한 2차 데이터 목록)

목적
- 검증사례/탄소발자국 산정 데이터.xlsx 의 4개 시트 레이아웃을 재현
- 토리컴(NiSO4) run05 결과 1,065.43 kgCO2e/ton 을 채워서 출력
- _original (OOXML) 의 디자인 발상은 참고하되, 셀/스타일은 우리 손으로 생성

실행
  python scripts/generate-calc-workbook-poc.py

출력
  검증사례/poc/toricomm_calc_workbook.xlsx
"""

from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ===========================================================================
# 데이터 모델
# ===========================================================================

@dataclass
class StageEmission:
    name: str
    value_kg_co2e: float


@dataclass
class SubCategoryEmission:
    name: str  # Biogenic / Fossil / LUC
    value_kg_co2e: float


@dataclass
class ProductCFP:
    code: str
    display_name: str
    functional_unit: str
    fu_label: str
    impact_category: str = "Climate change - global warming potential (GWP100)"
    unit: str = "kg CO2-Eq."
    total: float = 0.0
    stages: list[StageEmission] = field(default_factory=list)
    subcategories: list[SubCategoryEmission] = field(default_factory=list)


@dataclass
class StudyMeta:
    """표지 시트용 — CFP 연구 메타데이터 (KS I ISO 14067 §6.3 목표 및 범위)"""
    project_title: str
    client_company: str
    client_address: str
    contact_name: str
    contact_phone: str
    contact_email: str
    consultants: list[tuple[str, str, str]]  # (이름, 전화, 이메일)
    study_date: str
    standard: str
    gwp_basis: str
    assessment_period: str
    purpose: str


@dataclass
class SensitivityScenario:
    name: str
    baseline: str
    delta_kg_co2e: float
    delta_pct: float
    in_range: bool
    note: str = ""


@dataclass
class SecondaryDataItem:
    """KS I ISO 14067 §7.3 d) 데이터원 + §6.3.5 데이터 품질"""
    seq: int
    owner: str            # Data set owner (예: thinkstep AG, ecoinvent, 한국환경공단)
    uuid: str             # Activity/Product UUID (없으면 빈 문자열)
    activity_name: str
    geography: str        # KR / GLO / RoW 등
    reference_product: str
    unit: str
    amount: float
    ef_kg_co2e: float
    note: str = ""        # 비고 (전기, 운송, 원료물질 등 카테고리)


# ---------------------------------------------------------------------------
# 제품별 CFP 시트용 BOM 항목 (sheet5 'E2B-03E' 레이아웃 단순화 차용)
# ---------------------------------------------------------------------------

@dataclass
class BomItem:
    """KS I ISO 14067 §6.4.2 데이터 수집 + §6.3.5 데이터 품질"""
    direction: str          # 'input' | 'output'
    category: str           # 원료물질 / 보조물질 / 에너지 / 유틸리티 / 육상운송 / 제품 / 매립 / 폐수 / 포장 등
    name: str
    collected_unit: str
    collected_qty: float
    applied_unit: str       # 보통 collected_unit 과 동일
    applied_qty: float
    cut_off: str = ""       # cut-off 적용 시 사유, 그 외 빈 문자열
    activity_name: str = ""
    flow_name: str = ""
    location: str = ""
    ef_kg_co2e_per_unit: float = 0.0
    transport_mode: str = ""        # '국내육상운송' 등 (운송 항목에만)
    transport_distance_km: float = 0.0
    note: str = ""
    # DQR (1=best ~ 5=worst), KS I ISO 14067 §6.3.5 a~c
    dqr_ter: int = 3        # Time
    dqr_ger: int = 3        # Geographic
    dqr_tir: int = 3        # Technology
    # 농도 환산 — H2SO4(98%, 시판 EF 기준)는 적용 X, NaOH/H2O2는 적용 O
    concentration_pct: float = 100.0
    apply_concentration: bool = False
    # EF DB 매핑 — '사용한 2차 데이터 목록' 시트의 seq (1-base, 0이면 미매핑/cut-off)
    ef_seq: int = 0
    # 월별 분해 (KS I ISO 14067 §6.3.6 — 시간 경계, 12개월 변동성 추적)
    # None이면 단일값 모드 (collected_qty 만 사용), 12개 list 면 월별 + 합계 SUM 모드
    collected_monthly: list[float] | None = None
    # 전력 항목 메타 (KS I ISO 14067 §6.4.9.4 — 전력 처리 구분 의무)
    power_source_type: str = ""    # '외부그리드' | '자체발전' | '직접연결' | 'REC인증재생'
    power_supplier: str = ""       # 공급자명 (예: '한국전력공사 — 한국 평균')
    # 폐기물 항목 메타 (KS I ISO 14067 §6.3.8 — 폐기 시나리오 문서화)
    treatment_method: str = ""     # '재활용' | '매립' | '소각' | '위탁처리' | '폐수처리'
    treatment_facility: str = ""   # 처리업체명
    treatment_distance_km: float = 0.0  # 처리시설까지 운송거리


# ===========================================================================
# 토리컴 run05 데이터
# ===========================================================================

TORICOMM_NISO4 = ProductCFP(
    code="NiSO4-99.99-Granule",
    display_name="황산니켈 (NiSO4 99.99% Granule 0.5~1.5mm)",
    functional_unit="1 ton NiSO4 (FIBC 1Ton/Bag 출하 기준)",
    fu_label="황산니켈 1 ton",
    total=1065.43,
    stages=[
        StageEmission("원료 채취", 309.03),
        StageEmission("제조", 696.95),
        StageEmission("운송", 55.73),
        StageEmission("포장", 3.72),
    ],
    subcategories=[
        SubCategoryEmission("Climate change - Biogenic", 0.0),
        SubCategoryEmission("Climate change - Fossil", 1065.43),
        SubCategoryEmission("Climate change - Land use and land use change", 0.0),
    ],
)

TORICOMM_META = StudyMeta(
    project_title="(주)토리컴 황산니켈 제품 탄소발자국 산정",
    client_company="(주)토리컴",
    client_address="충남 아산시 (본사·생산공장)",
    contact_name="(미정 — 클라이언트 입력)",
    contact_phone="(미정 — 클라이언트 입력)",
    contact_email="(미정 — 클라이언트 입력)",
    consultants=[
        ("카보니 (Carbony) AI 컨설턴트", "—", "—"),
    ],
    study_date="2026-04-27 (run05, 인계 직전)",
    standard="ISO 14067:2018 / KS I ISO 14067",
    gwp_basis="IPCC AR6, 100년 (GWP100)",
    assessment_period="2025년 1월~12월 (12개월 평균)",
    purpose="양극재 고객사 공급망 CFP 보고 대응 (1차 자체 산정)",
)

TORICOMM_SENSITIVITY: list[SensitivityScenario] = [
    SensitivityScenario("전력 사용량 +20%", "980 kWh", +81.79, +0.077, True),
    SensitivityScenario("전력 사용량 -20%", "980 kWh", -81.79, -0.077, True),
    SensitivityScenario("NaOH 사용량 +20%", "380 kg (50% 용액)", +22.80, +0.021, True),
    SensitivityScenario("H2O2 사용량 +20%", "130 kg (35% 용액)", +13.65, +0.013, True),
    SensitivityScenario("운송 EF 상한 (0.13)", "0.10", +16.72, +0.016, True),
    SensitivityScenario("경제적 배분 (조황산니켈)", "데이터 부족", 0.0, 0.0, False, "확인 필요 — 매입 단가 확보 후 별도 분석"),
    SensitivityScenario("전력 EF 데이터셋 (0.4594)", "0.4173 (한국환경공단 2023)", +41.25, +0.039, True),
]

# EF DB 시트(사용한 2차 데이터 목록) 의 seq → BomItem.ef_seq 로 매핑되는 정렬된 EF 리스트.
# 정렬은 BOM 순서를 따른다 (검증 시 시각적 추적 용이).
TORICOMM_SECONDARY: list[SecondaryDataItem] = [
    SecondaryDataItem(1, "CarbonMate 내장 LCI DB", "", "황산 (98%) 생산",
                      "GLO", "Sulphuric acid", "kg", 1.0, 0.14, "원료물질"),
    SecondaryDataItem(2, "CarbonMate 내장 LCI DB", "", "수산화나트륨 50% 용액 생산",
                      "GLO", "Sodium hydroxide (50% sol.)", "kg", 1.0, 1.20, "원료물질 (NaOH 농도 환산 적용)"),
    SecondaryDataItem(3, "CarbonMate 내장 LCI DB", "", "과산화수소 35% 용액 생산",
                      "GLO", "Hydrogen peroxide (35% sol.)", "kg", 1.0, 1.50, "원료물질 (H2O2 농도 환산 적용)"),
    SecondaryDataItem(4, "CarbonMate 내장 LCI DB", "", "공업용수 공급",
                      "KR", "Industrial water", "m3", 1.0, 0.35, "유틸리티"),
    SecondaryDataItem(5, "한국환경공단 (KECO)", "", "한국 전력 그리드 평균",
                      "KR", "전력 (저압)", "kWh", 1.0, 0.4173, "전기 (앱 내장 / 2023년 기준)"),
    SecondaryDataItem(6, "CarbonMate 내장 LCI DB", "", "LNG 직접연소 (산업용)",
                      "KR", "Natural gas combustion", "Nm3", 1.0, 2.75, "연료"),
    SecondaryDataItem(7, "CarbonMate 내장 LCI DB", "", "산업용 외부 구매 스팀",
                      "KR", "Industrial steam, purchased", "kg", 1.0, 0.22, "에너지 (외부 구매 가정)"),
    SecondaryDataItem(8, "CarbonMate 내장 LCI DB", "", "16-32t 디젤 화물차량, ton-km 평균",
                      "KR", "Articulated lorry transport, 16-32t", "ton-km", 1.0, 0.10, "육상운송 (입고/출하 공통)"),
    SecondaryDataItem(9, "CarbonMate 내장 LCI DB", "", "PP (폴리프로필렌) 생산",
                      "GLO", "Polypropylene (PP) granulate", "kg", 1.0, 1.86, "포장 (FIBC 빅백)"),
    SecondaryDataItem(10, "CarbonMate 내장 LCI DB", "", "무기성 슬러지 매립",
                      "KR", "Landfill, inert waste", "kg", 1.0, 0.03, "폐기물 (중화 슬러지)"),
    SecondaryDataItem(11, "CarbonMate 내장 LCI DB", "", "지정폐기물 처리",
                      "KR", "Hazardous waste treatment", "kg", 1.0, 1.20, "폐기물 (지정폐기물)"),
    SecondaryDataItem(12, "CarbonMate 내장 LCI DB", "", "산업폐수 처리",
                      "KR", "Industrial wastewater treatment", "m3", 1.0, 0.40, "폐기물"),
]


# BOM — ef_seq 는 위 TORICOMM_SECONDARY 의 seq 와 매칭 (cross-sheet ref 자동 생성용)
TORICOMM_BOM: list[BomItem] = [
    # --- INPUT: 원료물질 (cut-off 2건 + EF 적용 3건) ---
    BomItem("input", "원료물질", "조황산니켈 (구리 정련 부산물)",
            "kg", 1450.0, "kg", 1450.0,
            cut_off="Cut-off — 구리 정련의 결정 제품 외 부산물 (ISO 14044 §5.3.5 ①)",
            location="KR", dqr_ter=1, dqr_ger=1, dqr_tir=1, ef_seq=0),
    BomItem("input", "원료물질", "배터리 슬러지 (폐기물 유래)",
            "kg", 620.0, "kg", 620.0,
            cut_off="Cut-off — 발생자가 폐기물로 인계 (ISO 14067 zero-burden)",
            location="KR", dqr_ter=1, dqr_ger=1, dqr_tir=1, ef_seq=0),
    BomItem("input", "원료물질", "황산 (H2SO4 98%, 시판품)",
            "kg", 60.0, "kg", 60.0,
            note="EF은 시판 농도 98% 기준 — 농도 환산 미적용",
            dqr_ter=3, dqr_ger=4, dqr_tir=2,
            concentration_pct=98.0, apply_concentration=False, ef_seq=1),
    BomItem("input", "원료물질", "수산화나트륨 (NaOH 50% 용액)",
            "kg", 380.0, "kg", 190.0,
            note="원액 380 kg × 농도 50% = 순물질 190 kg",
            dqr_ter=3, dqr_ger=4, dqr_tir=2,
            concentration_pct=50.0, apply_concentration=True, ef_seq=2),
    BomItem("input", "원료물질", "과산화수소 (H2O2 35% 용액)",
            "kg", 130.0, "kg", 45.5,
            note="원액 130 kg × 농도 35% = 순물질 45.5 kg",
            dqr_ter=3, dqr_ger=4, dqr_tir=2,
            concentration_pct=35.0, apply_concentration=True, ef_seq=3),
    # --- INPUT: 유틸리티 ---
    BomItem("input", "유틸리티", "공업용수 (용해 8.5 + 정제 4.0 m³)",
            "m3", 12.5, "m3", 12.5,
            dqr_ter=2, dqr_ger=1, dqr_tir=2, ef_seq=4),
    # --- INPUT: 에너지 ---
    BomItem("input", "에너지", "한국 전력 (용해220+중화95+정제180+결정화410+제품화75)",
            "kWh", 980.0, "kWh", 980.0,
            note="단계별 합산 — 용해+중화+정제+결정화+제품화 = 980 kWh",
            dqr_ter=2, dqr_ger=1, dqr_tir=1, ef_seq=5,
            power_source_type="외부그리드",
            power_supplier="한국전력공사 — 한국 평균 (KECO 2023)"),
    BomItem("input", "에너지", "LNG 직접연소 (보일러)",
            "Nm3", 12.0, "Nm3", 12.0,
            dqr_ter=2, dqr_ger=1, dqr_tir=2, ef_seq=6),
    BomItem("input", "에너지", "산업용 외부 구매 스팀",
            "kg", 850.0, "kg", 850.0,
            note="외부 구매 가정 — 자체 보일러 시 LNG 별도 적용 필요",
            dqr_ter=3, dqr_ger=1, dqr_tir=3, ef_seq=7),
    # --- INPUT: 운송 (모두 ef_seq=8 으로 매핑 — 동일 EF 0.10) ---
    BomItem("input", "육상운송", "조황산니켈 입고 (울산→아산)",
            "ton-km", 391.5, "ton-km", 391.5,
            transport_mode="국내육상운송", transport_distance_km=270.0,
            note="1.45 t × 270 km",
            dqr_ter=2, dqr_ger=1, dqr_tir=2, ef_seq=8),
    BomItem("input", "육상운송", "배터리 슬러지 입고 (충북→아산)",
            "ton-km", 55.8, "ton-km", 55.8,
            transport_mode="국내육상운송", transport_distance_km=90.0,
            note="0.62 t × 90 km",
            dqr_ter=2, dqr_ger=1, dqr_tir=2, ef_seq=8),
    BomItem("input", "육상운송", "황산니켈 출하 (아산→충북 고객사)",
            "ton-km", 110.0, "ton-km", 110.0,
            transport_mode="국내육상운송", transport_distance_km=110.0,
            note="1.0 t × 110 km",
            dqr_ter=2, dqr_ger=1, dqr_tir=2, ef_seq=8),
    # --- INPUT: 포장재 ---
    BomItem("input", "포장", "FIBC 빅백 (PP) 1 EA",
            "kg", 2.0, "kg", 2.0,
            note="앱 내장 1.86 vs 데이터셋 2.00 — 차이 5% 이내",
            dqr_ter=3, dqr_ger=4, dqr_tir=2, ef_seq=9),
    # --- OUTPUT: 제품 (FU anchor) ---
    BomItem("output", "제품", "황산니켈 (NiSO4 99.99% Granule)",
            "kg", 1000.0, "kg", 1000.0,
            note="기능단위 anchor — 모든 N열 환산의 분모",
            dqr_ter=1, dqr_ger=1, dqr_tir=1, ef_seq=0),
    # --- OUTPUT: 폐기물 ---
    BomItem("output", "매립", "중화 슬러지 (무기성)",
            "kg", 320.0, "kg", 320.0,
            dqr_ter=2, dqr_ger=1, dqr_tir=3, ef_seq=10,
            treatment_method="매립 (안정화)",
            treatment_facility="(미정 — 클라이언트 입력)",
            treatment_distance_km=45.0),
    BomItem("output", "지정폐기물", "지정폐기물 처리",
            "kg", 45.0, "kg", 45.0,
            dqr_ter=2, dqr_ger=1, dqr_tir=3, ef_seq=11,
            treatment_method="위탁처리",
            treatment_facility="(미정 — 클라이언트 입력)",
            treatment_distance_km=80.0),
    BomItem("output", "폐수", "산업폐수 처리",
            "m3", 11.0, "m3", 11.0,
            dqr_ter=2, dqr_ger=1, dqr_tir=3, ef_seq=12,
            treatment_method="폐수처리 (자체+위탁)",
            treatment_facility="(미정 — 클라이언트 입력)",
            treatment_distance_km=0.0),
]


# ===========================================================================
# 2번째 제품 (가상) — 다제품 아키텍처 검증용
# ===========================================================================
# ⚠️ 본 데이터는 검증되지 않은 가상 시나리오로, 실제 토리컴 산정값이 아님.
# 다제품 워크북 구조의 동작 확인 목적으로만 사용.
# 가정: 동일 공장 내 NiSO4 Powder Grade — 입력 70% 스케일, 다른 운송경로
# ===========================================================================

TORICOMM_NISO4_POWDER_HYPO = ProductCFP(
    code="NiSO4-99.9-Powder-HYPO",
    display_name="[가상] 황산니켈 (NiSO4 99.9% Powder)",
    functional_unit="1 ton NiSO4 Powder (가상 시나리오)",
    fu_label="황산니켈 Powder 1 ton",
    total=0.0,  # 미사용 — cross-ref 로 동적 계산
    stages=[
        StageEmission("원료 채취", 0.0),
        StageEmission("제조", 0.0),
        StageEmission("운송", 0.0),
        StageEmission("포장", 0.0),
    ],
    subcategories=[
        SubCategoryEmission("Climate change - Biogenic", 0.0),
        SubCategoryEmission("Climate change - Fossil", 0.0),
        SubCategoryEmission("Climate change - Land use and land use change", 0.0),
    ],
)


def _scale_bom(bom: list[BomItem], factor: float) -> list[BomItem]:
    """원본 BOM 의 모든 입력/출력 양을 factor 배 스케일한 사본을 만든다.
    cut-off 항목은 그대로 유지 (질량 비례) — 다제품 아키텍처 검증용.
    """
    from dataclasses import replace
    scaled: list[BomItem] = []
    for it in bom:
        scaled.append(replace(
            it,
            collected_qty=round(it.collected_qty * factor, 4),
            applied_qty=round(it.applied_qty * factor, 4),
            collected_monthly=(
                [round(v * factor, 4) for v in it.collected_monthly]
                if it.collected_monthly else None
            ),
        ))
    return scaled


# ---------------------------------------------------------------------------
# 가상 월별 분해 — 연 합계는 보존, 변동성·계절성 시뮬레이션
# ---------------------------------------------------------------------------
# ISO 14067 §6.3.6 (데이터 시간 경계) + §5.10 (정확성 — 변동성 정량 평가) 충족용.
# 토리컴 PoC 는 월별 raw 데이터 부재로 가상 분해를 사용하며, 본 사항은
# README 의 "알려진 한계" 항목에 명시된다.
# ---------------------------------------------------------------------------

# 12 개월 가중치 — 합계 12.0. 산업 평균 패턴 (여름철 6~8월 가동률 약 5% 상승 가정,
# 1·2월 설 연휴로 ~5% 감소)
_MONTHLY_WEIGHTS_NISO4 = [
    0.92, 0.92, 1.00, 1.04, 1.05, 1.06,
    1.05, 1.05, 1.02, 1.00, 0.95, 0.94,
]
# 합 12.0 검증
assert abs(sum(_MONTHLY_WEIGHTS_NISO4) - 12.0) < 1e-9, "월별 가중치 합은 12.0 이어야 함"


def _expand_monthly(annual_total: float, weights: list[float] | None = None) -> list[float]:
    """연 합계를 12개월로 분해. 가중치 미지정 시 토리컴 기본 패턴 사용.
    반환: 합계가 정확히 annual_total 이 되도록 마지막 월에서 보정된 12개 float.
    """
    w = weights if weights is not None else _MONTHLY_WEIGHTS_NISO4
    assert len(w) == 12, "weights 는 길이 12"
    monthly_avg = annual_total / 12.0
    raw = [round(monthly_avg * wi, 4) for wi in w]
    # 반올림 오차 보정 — 마지막 월에 차이 흡수
    diff = round(annual_total - sum(raw), 4)
    raw[-1] = round(raw[-1] + diff, 4)
    return raw


def _attach_monthly_to_bom(bom: list[BomItem], weights: list[float] | None = None) -> None:
    """BOM 모든 항목에 monthly 분해를 in-place 부착.
    cut-off / 제품 행 포함 — 모든 행이 monthly + SUM 패턴으로 통일됨.
    """
    for it in bom:
        if it.collected_monthly is None:
            it.collected_monthly = _expand_monthly(it.collected_qty, weights=weights)
        # 정합성 자동 검증 (반올림 오차 0.01 허용)
        delta = abs(sum(it.collected_monthly) - it.collected_qty)
        assert delta < 0.01 + abs(it.collected_qty) * 1e-4, (
            f"월별 합계 불일치: {it.name} monthly_sum={sum(it.collected_monthly)} vs "
            f"collected_qty={it.collected_qty} (delta={delta})"
        )


# Hypothetical product BOM — NiSO4 70% 스케일 + 운송거리 상이 (320km vs 270km 입고)
TORICOMM_BOM_POWDER_HYPO: list[BomItem] = _scale_bom(TORICOMM_BOM, factor=0.70)
# 운송 항목 1개만 거리 변경 — 다른 공급사 가정
# Original: 1.45 t × 270 km = 391.5 ton-km. After 0.70 scale → 1.015 t.
# HYPO override: 1.015 t × 320 km = 324.8 ton-km (입고)
for _it in TORICOMM_BOM_POWDER_HYPO:
    if _it.category == "육상운송" and "조황산니켈" in _it.name:
        _it.transport_distance_km = 320.0
        new_qty = round(1.45 * 0.70 * 320.0, 2)  # 324.8 ton-km
        _it.collected_qty = new_qty
        _it.applied_qty = new_qty
        _it.collected_monthly = None  # 강제 재계산 트리거
        _it.note = "[가상] 다른 공급사 — 1.015 t × 320 km = 324.8 ton-km"
        break

# 월별 분해 적용 (KS I ISO 14067 §6.3.6 — 시간 경계 충족)
# Powder HYPO 는 다른 가중치 패턴 사용 (다른 시장/계절성 가정) — 다양성 시연
_MONTHLY_WEIGHTS_POWDER = [
    1.10, 1.05, 1.00, 0.98, 0.95, 0.93,
    0.92, 0.95, 1.00, 1.05, 1.05, 1.02,
]
_attach_monthly_to_bom(TORICOMM_BOM, weights=_MONTHLY_WEIGHTS_NISO4)
_attach_monthly_to_bom(TORICOMM_BOM_POWDER_HYPO, weights=_MONTHLY_WEIGHTS_POWDER)


# ===========================================================================
# 스타일
# ===========================================================================

GRAY_FILL = PatternFill(start_color="FFBFBFBF", end_color="FFBFBFBF", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="FFE7E6E6", end_color="FFE7E6E6", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FF000000")
BODY_FONT = Font(name="맑은 고딕", size=10)
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True)
SUBTITLE_FONT = Font(name="맑은 고딕", size=11, bold=True)
NOTE_FONT = Font(name="맑은 고딕", size=9, color="FF595959")

THIN = Side(border_style="thin", color="FF808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)


def style_header(cell):
    cell.fill = GRAY_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER


def style_body(cell, *, numeric: bool = False, fmt: str = "#,##0.00###"):
    cell.fill = WHITE_FILL
    cell.font = BODY_FONT
    cell.alignment = CENTER if numeric else LEFT
    cell.border = BORDER
    if numeric:
        cell.number_format = fmt


# ===========================================================================
# 시트 1 — 표지
# ===========================================================================

def build_cover(ws, m: StudyMeta) -> None:
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 60

    # 제목
    ws.cell(row=2, column=2, value=m.project_title).font = TITLE_FONT
    ws.merge_cells("B2:D2")
    ws.cell(row=2, column=2).alignment = CENTER

    # 보조 정보 — 연구 범위 (ISO 14067 §6.3.2)
    rows = [
        ("연구 정보", "산정일", m.study_date),
        (None, "표준", m.standard),
        (None, "GWP 기준", m.gwp_basis),
        (None, "산정 대상 기간", m.assessment_period),
        (None, "산정 목적", m.purpose),
        ("의뢰사 (Client)", "회사명", m.client_company),
        (None, "소재지", m.client_address),
        (None, "담당자 이름", m.contact_name),
        (None, "전화번호", m.contact_phone),
        (None, "E-mail", m.contact_email),
    ]

    r = 4
    group_start = r
    last_group: str | None = None
    for group, label, value in rows:
        if group is not None and group != last_group:
            if last_group is not None and group_start < r:
                # 이전 그룹의 첫 번째 컬럼 머지
                ws.merge_cells(start_row=group_start, start_column=2, end_row=r - 1, end_column=2)
            group_start = r
            last_group = group
            ws.cell(row=r, column=2, value=group)
            style_header(ws.cell(row=r, column=2))
        elif group is None:
            # 그룹 셀은 비워두고 머지로 채움
            ws.cell(row=r, column=2)
            style_header(ws.cell(row=r, column=2))

        ws.cell(row=r, column=3, value=label); style_header(ws.cell(row=r, column=3))
        ws.cell(row=r, column=4, value=value); style_body(ws.cell(row=r, column=4))
        r += 1

    # 마지막 그룹 머지
    if last_group is not None and group_start < r:
        ws.merge_cells(start_row=group_start, start_column=2, end_row=r - 1, end_column=2)

    # 수행자(컨설턴트) 블록
    r += 1
    ws.cell(row=r, column=2, value="수행자 (Consultants)").font = SUBTITLE_FONT
    r += 1
    for col, label in enumerate(["순번", "이름", "연락처"], start=2):
        style_header(ws.cell(row=r, column=col, value=label))
    # 연락처 컬럼은 D 한 칸 사용
    r += 1
    for i, (name, phone, email) in enumerate(m.consultants, start=1):
        ws.cell(row=r, column=2, value=i); style_body(ws.cell(row=r, column=2), numeric=True, fmt="0")
        ws.cell(row=r, column=3, value=name); style_body(ws.cell(row=r, column=3))
        contact = " / ".join([x for x in (phone, email) if x and x != "—"]) or "—"
        ws.cell(row=r, column=4, value=contact); style_body(ws.cell(row=r, column=4))
        r += 1

    # 푸터 노트 (ISO 14067 §5.11 투명성 — 자가선언)
    r += 1
    note = ws.cell(
        row=r,
        column=2,
        value=("※ 본 산정 결과는 ISO 14067:2018 기반의 자체 산정이며, "
               "제3자 검증 통과를 보장하지 않습니다. "
               "검증 통과는 별도의 인증기관(LRQA, KFQ, KMR 등) 절차를 따릅니다."),
    )
    note.font = NOTE_FONT
    note.alignment = LEFT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)


# ===========================================================================
# 시트 2 — LCIA (기존 로직)
# ===========================================================================

def build_lcia_block(
    ws, *, start_row: int, p: ProductCFP,
    product_sheet_name: str | None = None,
    product_total_row: int | None = None,
    stage_row_map: dict[str, list[int]] | None = None,
) -> tuple[int, int, int]:
    """LCIA 블록 생성. 반환: (next_row, stage_first_row, stage_last_row)
    stage_first_row, stage_last_row 는 차트 데이터 참조 범위로 사용된다.
    """
    """LCIA 블록.
    product_sheet_name + product_total_row + stage_row_map 가 주어지면
    값을 cross-sheet 수식으로 참조 (× 1000 → 1 ton 단위 환산).
    그렇지 않으면 ProductCFP 정적 값 사용.
    """
    use_refs = product_sheet_name and product_total_row and stage_row_map

    r = start_row
    ws.cell(row=r, column=2, value=f"{p.display_name}의 탄소발자국").font = SUBTITLE_FONT
    r += 1
    ws.cell(row=r, column=2, value="영향범주"); style_header(ws.cell(row=r, column=2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    ws.cell(row=r, column=2, value=p.impact_category); style_body(ws.cell(row=r, column=2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1

    for col, label in enumerate(["제품명", "단위", "탄소발자국 값"], start=2):
        style_header(ws.cell(row=r, column=col, value=label))
    r += 1
    ws.cell(row=r, column=2, value=p.fu_label); style_body(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3, value=p.unit); style_body(ws.cell(row=r, column=3), numeric=True)
    if use_refs:
        # S{total_row} 는 kgCO2e/kg FU. ton 환산 × 1000.
        ws.cell(row=r, column=4, value=f"='{product_sheet_name}'!S{product_total_row}*1000")
    else:
        ws.cell(row=r, column=4, value=p.total)
    style_body(ws.cell(row=r, column=4), numeric=True)
    r += 1

    for col, label in enumerate(["전과정 단계", "단위", "탄소발자국 값"], start=2):
        style_header(ws.cell(row=r, column=col, value=label))
    r += 1
    stage_first_row = r
    for stage in p.stages:
        ws.cell(row=r, column=2, value=stage.name); style_body(ws.cell(row=r, column=2))
        ws.cell(row=r, column=3, value=p.unit); style_body(ws.cell(row=r, column=3), numeric=True)
        if use_refs and stage.name in stage_row_map:
            rows = stage_row_map[stage.name]
            sum_terms = ",".join(f"'{product_sheet_name}'!S{rr}" for rr in rows)
            ws.cell(row=r, column=4, value=f"=SUM({sum_terms})*1000")
        else:
            ws.cell(row=r, column=4, value=stage.value_kg_co2e)
        style_body(ws.cell(row=r, column=4), numeric=True)
        r += 1
    stage_last_row = r - 1

    for col, label in enumerate(["탄소발자국 하위범주", "단위", "탄소발자국 값"], start=2):
        style_header(ws.cell(row=r, column=col, value=label))
    r += 1
    for sub in p.subcategories:
        ws.cell(row=r, column=2, value=sub.name); style_body(ws.cell(row=r, column=2))
        ws.cell(row=r, column=3, value=p.unit); style_body(ws.cell(row=r, column=3), numeric=True)
        # Toricomm 무기화학: Biogenic=0, LUC=0, Fossil=합계 — 모두 합계 참조 또는 0
        if use_refs:
            if "Fossil" in sub.name:
                ws.cell(row=r, column=4, value=f"='{product_sheet_name}'!S{product_total_row}*1000")
            else:
                ws.cell(row=r, column=4, value=0)
        else:
            ws.cell(row=r, column=4, value=sub.value_kg_co2e)
        style_body(ws.cell(row=r, column=4), numeric=True)
        r += 1

    return r + 1, stage_first_row, stage_last_row


def build_lcia(
    ws,
    products: list[ProductCFP],
    *,
    product_sheet_names: dict[str, str] | None = None,
    product_total_rows: dict[str, int] | None = None,
    stage_row_maps: dict[str, dict[str, list[int]]] | None = None,
) -> None:
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18

    next_row = 2
    for p in products:
        kwargs = {}
        if product_sheet_names and p.code in product_sheet_names:
            kwargs["product_sheet_name"] = product_sheet_names[p.code]
            kwargs["product_total_row"] = product_total_rows[p.code]
            kwargs["stage_row_map"] = stage_row_maps[p.code]
        block_top = next_row
        next_row, stage_first, stage_last = build_lcia_block(
            ws, start_row=next_row, p=p, **kwargs
        )

        # G — 단계별 기여도 파이 차트 (살아있는 셀 참조)
        # 차트는 LCIA 블록의 오른쪽(F열)에 배치되어 같은 행 범위로 정렬됨.
        chart = PieChart()
        chart.title = f"{p.display_name} — 전과정 단계별 기여도"
        labels = Reference(ws, min_col=2, min_row=stage_first, max_row=stage_last)
        data = Reference(ws, min_col=4, min_row=stage_first - 1, max_row=stage_last)  # min_row -1 = 헤더 (시리즈명)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.dataLabels = DataLabelList(showPercent=True, showCatName=True)
        chart.height = 9   # cm
        chart.width = 14   # cm
        ws.add_chart(chart, f"F{block_top}")


# ===========================================================================
# 시트 3 — 민감도 분석 (KS I ISO 14067 §6.4.6.1, §6.6 — 민감도 분석 필수)
# ===========================================================================

def build_sensitivity(ws, *, product: ProductCFP, scenarios: list[SensitivityScenario]) -> None:
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 36

    # 제목
    ws.cell(row=1, column=2, value="민감도 분석").font = TITLE_FONT

    # 서술 (5.10 정확성 + 5.11 투명성 — 가정의 영향 정량 분석)
    intro_lines = [
        f"ㅇ 본 분석은 {product.display_name} ({product.fu_label})의 CFP "
        f"{product.total:,.2f} {product.unit} 를 기준으로 한다.",
        "ㅇ 주요 활동량/EF/방법론 변경이 최종 결과에 미치는 영향을 정량 평가한다.",
        "ㅇ 변화율 ±10% 이내를 '범위 내'로 판단하며, 그 외는 추가 분석 또는 개선 권고 대상이다.",
        "ㅇ ISO 14067 §6.4.6.1 (대체 할당 절차 적용 시 민감도 분석 의무) 및 §6.6 (해석 단계 민감도 분석 의무) 준수.",
    ]
    for i, line in enumerate(intro_lines):
        c = ws.cell(row=2 + i, column=2, value=line)
        c.font = BODY_FONT
        c.alignment = LEFT
        ws.merge_cells(start_row=2 + i, start_column=2, end_row=2 + i, end_column=7)

    # 영향범주 헤더
    r = 7
    ws.cell(row=r, column=2, value="영향범주"); style_header(ws.cell(row=r, column=2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 1
    ws.cell(row=r, column=2, value=product.impact_category); style_body(ws.cell(row=r, column=2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 2

    # 시나리오 표 헤더
    headers = ["시나리오", "기준값", "변화량 (kgCO₂e)", "변화율", "범위 내", "비고"]
    for col, h in enumerate(headers, start=2):
        style_header(ws.cell(row=r, column=col, value=h))
    r += 1

    # 시나리오 행 + 합산용 baseline (수식)
    baseline_row = r
    ws.cell(row=baseline_row, column=2, value=f"기준값 ({product.fu_label} CFP)")
    style_body(ws.cell(row=baseline_row, column=2))
    ws.cell(row=baseline_row, column=3, value=product.unit)
    style_body(ws.cell(row=baseline_row, column=3))
    ws.cell(row=baseline_row, column=4, value=product.total)
    style_body(ws.cell(row=baseline_row, column=4), numeric=True)
    ws.cell(row=baseline_row, column=5, value="—"); style_body(ws.cell(row=baseline_row, column=5), numeric=True)
    ws.cell(row=baseline_row, column=6, value="—"); style_body(ws.cell(row=baseline_row, column=6), numeric=True)
    ws.cell(row=baseline_row, column=7, value="기준 (Baseline)"); style_body(ws.cell(row=baseline_row, column=7))
    r += 1

    scenarios_first_row = r
    for s in scenarios:
        ws.cell(row=r, column=2, value=s.name); style_body(ws.cell(row=r, column=2))
        ws.cell(row=r, column=3, value=s.baseline); style_body(ws.cell(row=r, column=3))
        ws.cell(row=r, column=4, value=s.delta_kg_co2e); style_body(ws.cell(row=r, column=4), numeric=True)
        ws.cell(row=r, column=5, value=s.delta_pct); style_body(ws.cell(row=r, column=5), numeric=True, fmt="0.0%")
        mark = "✅" if s.in_range else "⚠️"
        ws.cell(row=r, column=6, value=mark); style_body(ws.cell(row=r, column=6), numeric=True)
        ws.cell(row=r, column=7, value=s.note or ""); style_body(ws.cell(row=r, column=7))
        if not s.in_range:
            for col in range(2, 8):
                ws.cell(row=r, column=col).fill = WARN_FILL
        r += 1
    scenarios_last_row = r - 1

    # G — 시나리오 변화율 막대 차트 (살아있는 셀 참조: B=라벨, E=변화율)
    bar = BarChart()
    bar.type = "bar"
    bar.style = 11
    bar.title = "민감도 분석 — 시나리오별 CFP 변화율"
    bar.y_axis.title = None
    bar.x_axis.title = "변화율"
    labels = Reference(ws, min_col=2, min_row=scenarios_first_row, max_row=scenarios_last_row)
    data = Reference(ws, min_col=5, min_row=scenarios_first_row - 1, max_row=scenarios_last_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.dataLabels = DataLabelList(showVal=True)
    bar.height = 9
    bar.width = 18
    ws.add_chart(bar, f"I{scenarios_first_row - 2}")

    # 결론
    r += 1
    conclusion = (
        "ㅇ 결론: 전력이 CFP의 가장 큰 단일 변동 인자(±7.7%)이며, 그 외 항목은 모두 ±5% 이내. "
        "조황산니켈의 경제적 배분은 매입 단가 확보 후 별도 분석이 필요하다 (ISO 14044 §5.3.5)."
    )
    c = ws.cell(row=r, column=2, value=conclusion)
    c.font = BODY_FONT; c.alignment = LEFT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)


# ===========================================================================
# 시트 4 — 사용한 2차 데이터 목록 (KS I ISO 14067 §7.3 d) 데이터원)
# ===========================================================================

def build_secondary_data(ws, items: list[SecondaryDataItem]) -> None:
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 6   # 순번
    ws.column_dimensions["C"].width = 22  # owner
    ws.column_dimensions["D"].width = 24  # uuid
    ws.column_dimensions["E"].width = 38  # activity name
    ws.column_dimensions["F"].width = 10  # geography
    ws.column_dimensions["G"].width = 32  # ref product
    ws.column_dimensions["H"].width = 8   # unit
    ws.column_dimensions["I"].width = 10  # amount
    ws.column_dimensions["J"].width = 12  # ef
    ws.column_dimensions["K"].width = 28  # 비고

    # 메타 (LCIA 메서드 정보)
    meta_pairs = [
        ("Method", "IPCC AR6 / KS I ISO 14067:2018"),
        ("Category", "climate change"),
        ("Indicator", "global warming potential (GWP100)"),
    ]
    for i, (k, v) in enumerate(meta_pairs):
        rr = 2 + i
        ws.cell(row=rr, column=10, value=k); style_header(ws.cell(row=rr, column=10))
        ws.cell(row=rr, column=11, value=v); style_body(ws.cell(row=rr, column=11))

    # 인트로 노트 (ISO 14067 §6.3.5 — 2차 데이터 정당화·문서화 의무)
    ws.cell(row=2, column=4, value="ㅇ 1차 데이터 수집이 실행 가능하지 않은 투입물·산출물에 한해 2차 데이터를 사용함.").font = BODY_FONT
    ws.cell(row=3, column=4, value="ㅇ 모든 2차 데이터는 출처(Owner)·UUID·지리적 범위를 명시하여 추적 가능성을 확보함.").font = BODY_FONT
    ws.cell(row=4, column=4, value="ㅇ ISO 14067 §6.3.5 (데이터 품질) 및 §7.3 d) (데이터원 기록 의무) 준수.").font = BODY_FONT

    # 헤더
    r = 7
    headers = [
        "순번", "Data set owner", "Activity UUID / Product UUID",
        "Activity Name", "Geography", "Reference Product Name",
        "Unit", "Amount", "kg CO₂-Eq", "비고",
    ]
    for col, h in enumerate(headers, start=2):
        style_header(ws.cell(row=r, column=col, value=h))
    r += 1

    for it in items:
        ws.cell(row=r, column=2, value=it.seq); style_body(ws.cell(row=r, column=2), numeric=True, fmt="0")
        ws.cell(row=r, column=3, value=it.owner); style_body(ws.cell(row=r, column=3))
        ws.cell(row=r, column=4, value=it.uuid or "(N/A — 내장 DB)"); style_body(ws.cell(row=r, column=4))
        ws.cell(row=r, column=5, value=it.activity_name); style_body(ws.cell(row=r, column=5))
        ws.cell(row=r, column=6, value=it.geography); style_body(ws.cell(row=r, column=6), numeric=True)
        ws.cell(row=r, column=7, value=it.reference_product); style_body(ws.cell(row=r, column=7))
        ws.cell(row=r, column=8, value=it.unit); style_body(ws.cell(row=r, column=8), numeric=True)
        ws.cell(row=r, column=9, value=it.amount); style_body(ws.cell(row=r, column=9), numeric=True)
        ws.cell(row=r, column=10, value=it.ef_kg_co2e); style_body(ws.cell(row=r, column=10), numeric=True, fmt="#,##0.0000")
        ws.cell(row=r, column=11, value=it.note); style_body(ws.cell(row=r, column=11))
        r += 1

    # 푸터 노트
    r += 1
    c = ws.cell(
        row=r, column=2,
        value=("※ 한국환경공단(KECO) 전력 EF는 매년 갱신됨. "
               "검증 시점에 따라 가장 최근 공표값을 적용해야 한다 (ISO 14067 §6.3.6 시간 경계 / §6.4.9.4 전력)."),
    )
    c.font = NOTE_FONT; c.alignment = LEFT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)


# ===========================================================================
# 시트 (NEW) — 제품 생산량
# ===========================================================================
# 원본 '제품 생산량' 시트는 12개월 컬럼 + 합계 형태. 토리컴은 연 평균이므로
# 단일 합계 컬럼만 사용. row 위치: r3 = FU anchor (1000 kg = 1 ton)
# ===========================================================================

PRODUCTION_HEADER_ROW = 2
PRODUCTION_DATA_START = 3  # FU anchor 행


# 제품 생산량 — 12개월 + 합계 SUM
PROD_MONTH_FIRST_COL = 5    # E
PROD_MONTH_LAST_COL = 16    # P
PROD_SUM_COL = 17           # Q
PROD_FU_ANCHOR_LETTER = chr(64 + PROD_SUM_COL)  # "Q"


def build_production(
    ws,
    products: list[ProductCFP],
    fu_kg_by_product: dict[str, float],
    monthly_weights_by_product: dict[str, list[float]] | None = None,
) -> dict[str, int]:
    """제품 생산량 시트 (월별 12 컬럼 + 합계).
    KS I ISO 14067 §6.3.6 준수.
    반환: {product.code: fu_anchor_row}  — Q 열의 합계 셀이 FU 분모로 사용됨.
    """
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 8
    for c in range(PROD_MONTH_FIRST_COL, PROD_MONTH_LAST_COL + 1):
        ws.column_dimensions[chr(64 + c)].width = 9
    ws.column_dimensions[chr(64 + PROD_SUM_COL)].width = 13

    last_col = PROD_SUM_COL
    ws.cell(row=1, column=1,
            value="3. 제품 생산량 — 월별 12 컬럼 (FU 분모 = Q 열 합계)").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    HEADER_TOP = PRODUCTION_HEADER_ROW
    HEADER_SUB = HEADER_TOP + 1
    fixed_left = [(1, "순번"), (2, "제품구분"), (3, "제품명 (코드)"), (4, "단위")]
    for col, lbl in fixed_left:
        style_header(ws.cell(row=HEADER_TOP, column=col, value=lbl))
        ws.merge_cells(start_row=HEADER_TOP, start_column=col,
                       end_row=HEADER_SUB, end_column=col)
    style_header(ws.cell(row=HEADER_TOP, column=PROD_MONTH_FIRST_COL,
                         value="월별 생산량 (kg)"))
    ws.merge_cells(start_row=HEADER_TOP, start_column=PROD_MONTH_FIRST_COL,
                   end_row=HEADER_TOP, end_column=PROD_MONTH_LAST_COL)
    for i, mlabel in enumerate(KOREAN_MONTHS):
        style_header(ws.cell(row=HEADER_SUB,
                             column=PROD_MONTH_FIRST_COL + i, value=mlabel))
    style_header(ws.cell(row=HEADER_TOP, column=PROD_SUM_COL, value="합계 (=SUM, FU 기준)"))
    ws.merge_cells(start_row=HEADER_TOP, start_column=PROD_SUM_COL,
                   end_row=HEADER_SUB, end_column=PROD_SUM_COL)

    fu_rows: dict[str, int] = {}
    r = HEADER_SUB + 1
    sum_first = chr(64 + PROD_MONTH_FIRST_COL)
    sum_last = chr(64 + PROD_MONTH_LAST_COL)
    for i, p in enumerate(products):
        fu = fu_kg_by_product.get(p.code, 1000.0)
        weights = (monthly_weights_by_product or {}).get(p.code) or _MONTHLY_WEIGHTS_NISO4
        monthly = _expand_monthly(fu, weights=weights)
        ws.cell(row=r, column=1, value=i + 1); style_body(ws.cell(row=r, column=1), numeric=True, fmt="0")
        ws.cell(row=r, column=2, value="주제품"); style_body(ws.cell(row=r, column=2))
        ws.cell(row=r, column=3, value=f"{p.display_name}  [{p.code}]"); style_body(ws.cell(row=r, column=3))
        ws.cell(row=r, column=4, value="kg"); style_body(ws.cell(row=r, column=4))
        for j, mv in enumerate(monthly):
            c = ws.cell(row=r, column=PROD_MONTH_FIRST_COL + j, value=mv)
            style_body(c, numeric=True)
        ws.cell(row=r, column=PROD_SUM_COL, value=f"=SUM({sum_first}{r}:{sum_last}{r})")
        style_body(ws.cell(row=r, column=PROD_SUM_COL), numeric=True)
        ws.cell(row=r, column=PROD_SUM_COL).font = HEADER_FONT
        fu_rows[p.code] = r
        r += 1

    note = ws.cell(row=r + 1, column=1,
                   value=("※ Q 열 (합계) 가 각 제품 CFP 시트의 N (FU 환산) 수식 분모로 참조됩니다. "
                          "월별 셀을 사용자가 직접 수정해도 합계·CFP·차트 모두 자동 갱신됩니다 (살아있는 수식)."))
    note.font = NOTE_FONT; note.alignment = LEFT
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=last_col)

    return fu_rows


# ===========================================================================
# 시트 (NEW) — BOM 입력물
# ===========================================================================
# 원본 '원료물질, 보조물질 투입량' (87r×22c) 의 핵심 패턴을 단순화:
#   - 12개월 컬럼 → 합계 1컬럼
#   - 농도 처리 명시화 (concentration_pct / apply_concentration)
#   - 적용수량 = IF(apply_concentration, 합계 × 농도/100, 합계)
# ===========================================================================

BOM_INPUT_HEADER_ROW = 2
BOM_INPUT_DATA_START = 3
BOM_OUTPUT_HEADER_ROW = 2
BOM_OUTPUT_DATA_START = 3


# ------------------------------------------------------------------
# BOM 입력물 — 12개월 컬럼 레이아웃
#   A 순번 / B 적용 제품 / C 분류 / D 명칭 / E 단위
#   F~Q  : 1월~12월 (12 컬럼)
#   R    : 합계 = =SUM(F:Q)
#   S    : 농도 (%)
#   T    : 농도 적용 여부 (Y/N)
#   U    : 적용수량 = =IF(T="Y", R*S/100, R)
#   V    : 운송거리 (km)
#   W    : DQI
#   X    : 비고
# ------------------------------------------------------------------
BOM_IN_MONTH_FIRST_COL = 6   # F
BOM_IN_MONTH_LAST_COL = 17   # Q
BOM_IN_SUM_COL = 18          # R  (합계)
BOM_IN_CONC_COL = 19         # S
BOM_IN_APPLY_COL = 20        # T
BOM_IN_QTY_COL = 21          # U  (적용수량)
BOM_IN_DIST_COL = 22         # V  (운송거리)
BOM_IN_DQI_COL = 23          # W
BOM_IN_NOTE_COL = 24         # X

KOREAN_MONTHS = [f"{i}월" for i in range(1, 13)]


def build_bom_input(
    ws,
    boms_by_product: dict[str, list[BomItem]],
    products: list[ProductCFP],
) -> dict[str, dict[int, int]]:
    """다제품 + 12개월 컬럼 BOM 입력물 시트.
    반환: {product.code: {bom_index_in_full_list: row_number}}

    KS I ISO 14067 §6.3.6 (시간 경계) — 12개월 변동성 추적.
    합계는 monthly SUM 수식이므로 사용자가 어느 월이든 수정하면 합계·CFP 자동 갱신.
    """
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 18  # 적용 제품
    ws.column_dimensions["C"].width = 14  # 분류
    ws.column_dimensions["D"].width = 36  # 명칭
    ws.column_dimensions["E"].width = 8   # 단위
    # F~Q : 12개월 (compact)
    for c in range(BOM_IN_MONTH_FIRST_COL, BOM_IN_MONTH_LAST_COL + 1):
        ws.column_dimensions[chr(64 + c)].width = 9
    ws.column_dimensions[chr(64 + BOM_IN_SUM_COL)].width = 13   # R 합계
    ws.column_dimensions[chr(64 + BOM_IN_CONC_COL)].width = 9   # S
    ws.column_dimensions[chr(64 + BOM_IN_APPLY_COL)].width = 9  # T
    ws.column_dimensions[chr(64 + BOM_IN_QTY_COL)].width = 12   # U
    ws.column_dimensions[chr(64 + BOM_IN_DIST_COL)].width = 11  # V
    ws.column_dimensions[chr(64 + BOM_IN_DQI_COL)].width = 7    # W
    ws.column_dimensions[chr(64 + BOM_IN_NOTE_COL)].width = 32  # X

    last_col = BOM_IN_NOTE_COL
    ws.cell(row=1, column=1,
            value="4. BOM — 입력물 (원료/보조/유틸/에너지/운송/포장) — 월별 12 컬럼"
            ).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    # 헤더 — 2행 머지 구조 (월 그룹 머지)
    HEADER_TOP = BOM_INPUT_HEADER_ROW    # = 2
    HEADER_SUB = HEADER_TOP + 1          # = 3
    fixed_left_headers = [
        (1, "순번"), (2, "적용 제품"), (3, "분류"), (4, "명칭"), (5, "단위"),
    ]
    for col, lbl in fixed_left_headers:
        style_header(ws.cell(row=HEADER_TOP, column=col, value=lbl))
        ws.merge_cells(start_row=HEADER_TOP, start_column=col,
                       end_row=HEADER_SUB, end_column=col)
    # F~Q 그룹 헤더
    style_header(ws.cell(row=HEADER_TOP, column=BOM_IN_MONTH_FIRST_COL,
                         value="월별 사용량 (kg / m³ / kWh / Nm³ / ton-km)"))
    ws.merge_cells(start_row=HEADER_TOP, start_column=BOM_IN_MONTH_FIRST_COL,
                   end_row=HEADER_TOP, end_column=BOM_IN_MONTH_LAST_COL)
    for i, mlabel in enumerate(KOREAN_MONTHS):
        style_header(ws.cell(row=HEADER_SUB,
                             column=BOM_IN_MONTH_FIRST_COL + i, value=mlabel))
    # R~X
    for col, lbl in [
        (BOM_IN_SUM_COL, "합계 (=SUM)"),
        (BOM_IN_CONC_COL, "농도 (%)"),
        (BOM_IN_APPLY_COL, "농도 적용"),
        (BOM_IN_QTY_COL, "적용수량"),
        (BOM_IN_DIST_COL, "운송거리(km)"),
        (BOM_IN_DQI_COL, "DQI"),
        (BOM_IN_NOTE_COL, "비고"),
    ]:
        style_header(ws.cell(row=HEADER_TOP, column=col, value=lbl))
        ws.merge_cells(start_row=HEADER_TOP, start_column=col,
                       end_row=HEADER_SUB, end_column=col)

    # 데이터 시작 행을 헤더 머지 후로 조정
    result: dict[str, dict[int, int]] = {}
    r = HEADER_SUB + 1
    seq = 1
    sum_first_letter = chr(64 + BOM_IN_MONTH_FIRST_COL)   # F
    sum_last_letter = chr(64 + BOM_IN_MONTH_LAST_COL)     # Q
    for p in products:
        bom = boms_by_product[p.code]
        result[p.code] = {}
        # 제품 구분 행 (시각적 분리)
        ws.cell(row=r, column=1, value=f"▼ {p.code}")
        ws.cell(row=r, column=1).font = SUBTITLE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        ws.cell(row=r, column=1).fill = LIGHT_GRAY_FILL
        r += 1

        for full_idx, it in enumerate(bom):
            if it.direction != "input":
                continue
            # 전기 사용량은 별도 시트로 라우팅됨 — BOM 입력물에서 제외
            if _bom_item_sheet(it) != SHEET_BOM_INPUT:
                continue
            ws.cell(row=r, column=1, value=seq); style_body(ws.cell(row=r, column=1), numeric=True, fmt="0")
            ws.cell(row=r, column=2, value=p.code); style_body(ws.cell(row=r, column=2))
            ws.cell(row=r, column=3, value=it.category); style_body(ws.cell(row=r, column=3))
            ws.cell(row=r, column=4, value=it.name); style_body(ws.cell(row=r, column=4))
            ws.cell(row=r, column=5, value=it.collected_unit); style_body(ws.cell(row=r, column=5))
            # F~Q 월별 값
            monthly = it.collected_monthly or [it.collected_qty / 12.0] * 12
            for i, mv in enumerate(monthly):
                c = ws.cell(row=r, column=BOM_IN_MONTH_FIRST_COL + i, value=mv)
                style_body(c, numeric=True)
            # R 합계 (살아있는 SUM)
            ws.cell(row=r, column=BOM_IN_SUM_COL,
                    value=f"=SUM({sum_first_letter}{r}:{sum_last_letter}{r})")
            style_body(ws.cell(row=r, column=BOM_IN_SUM_COL), numeric=True)
            ws.cell(row=r, column=BOM_IN_SUM_COL).font = HEADER_FONT
            # S 농도
            ws.cell(row=r, column=BOM_IN_CONC_COL, value=it.concentration_pct)
            style_body(ws.cell(row=r, column=BOM_IN_CONC_COL), numeric=True, fmt="0.0")
            # T 농도 적용 여부
            ws.cell(row=r, column=BOM_IN_APPLY_COL, value="Y" if it.apply_concentration else "N")
            style_body(ws.cell(row=r, column=BOM_IN_APPLY_COL))
            # U 적용수량 = IF(T="Y", R*S/100, R)
            t_let = chr(64 + BOM_IN_APPLY_COL)
            r_let = chr(64 + BOM_IN_SUM_COL)
            s_let = chr(64 + BOM_IN_CONC_COL)
            ws.cell(row=r, column=BOM_IN_QTY_COL,
                    value=f'=IF({t_let}{r}="Y",{r_let}{r}*{s_let}{r}/100,{r_let}{r})')
            style_body(ws.cell(row=r, column=BOM_IN_QTY_COL), numeric=True)
            # V 운송거리
            if it.transport_distance_km:
                ws.cell(row=r, column=BOM_IN_DIST_COL, value=it.transport_distance_km)
                style_body(ws.cell(row=r, column=BOM_IN_DIST_COL), numeric=True)
            else:
                ws.cell(row=r, column=BOM_IN_DIST_COL, value="")
                style_body(ws.cell(row=r, column=BOM_IN_DIST_COL))
            # W DQI
            dqi = "M" if (it.dqr_ter <= 2 and it.dqr_ger <= 2) else "C"
            ws.cell(row=r, column=BOM_IN_DQI_COL, value=dqi)
            style_body(ws.cell(row=r, column=BOM_IN_DQI_COL))
            # X 비고
            note_text = it.cut_off if it.cut_off else (it.note or "")
            ws.cell(row=r, column=BOM_IN_NOTE_COL, value=note_text)
            style_body(ws.cell(row=r, column=BOM_IN_NOTE_COL))
            if it.cut_off:
                for col in range(1, last_col + 1):
                    ws.cell(row=r, column=col).fill = WARN_FILL
            result[p.code][full_idx] = r
            r += 1
            seq += 1

    return result


# ===========================================================================
# 시트 (NEW) — BOM 출력물 (제품 + 폐기물)
# ===========================================================================

# BOM 출력물 — 입력물과 동일한 12개월 + SUM 패턴
BOM_OUT_MONTH_FIRST_COL = 6   # F
BOM_OUT_MONTH_LAST_COL = 17   # Q
BOM_OUT_SUM_COL = 18          # R
BOM_OUT_DQI_COL = 19          # S
BOM_OUT_NOTE_COL = 20         # T


def build_bom_output(
    ws,
    boms_by_product: dict[str, list[BomItem]],
    products: list[ProductCFP],
) -> dict[str, dict[int, int]]:
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 18  # 적용 제품
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 8
    for c in range(BOM_OUT_MONTH_FIRST_COL, BOM_OUT_MONTH_LAST_COL + 1):
        ws.column_dimensions[chr(64 + c)].width = 9
    ws.column_dimensions[chr(64 + BOM_OUT_SUM_COL)].width = 12
    ws.column_dimensions[chr(64 + BOM_OUT_DQI_COL)].width = 7
    ws.column_dimensions[chr(64 + BOM_OUT_NOTE_COL)].width = 32

    last_col = BOM_OUT_NOTE_COL
    ws.cell(row=1, column=1,
            value="5. BOM — 출력물 (제품 + 폐기물) — 월별 12 컬럼"
            ).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    HEADER_TOP = BOM_OUTPUT_HEADER_ROW   # 2
    HEADER_SUB = HEADER_TOP + 1          # 3
    fixed_left_headers = [
        (1, "순번"), (2, "적용 제품"), (3, "분류"), (4, "명칭"), (5, "단위"),
    ]
    for col, lbl in fixed_left_headers:
        style_header(ws.cell(row=HEADER_TOP, column=col, value=lbl))
        ws.merge_cells(start_row=HEADER_TOP, start_column=col,
                       end_row=HEADER_SUB, end_column=col)
    style_header(ws.cell(row=HEADER_TOP, column=BOM_OUT_MONTH_FIRST_COL,
                         value="월별 발생량 (kg / m³)"))
    ws.merge_cells(start_row=HEADER_TOP, start_column=BOM_OUT_MONTH_FIRST_COL,
                   end_row=HEADER_TOP, end_column=BOM_OUT_MONTH_LAST_COL)
    for i, mlabel in enumerate(KOREAN_MONTHS):
        style_header(ws.cell(row=HEADER_SUB,
                             column=BOM_OUT_MONTH_FIRST_COL + i, value=mlabel))
    for col, lbl in [
        (BOM_OUT_SUM_COL, "합계 (=SUM)"),
        (BOM_OUT_DQI_COL, "DQI"),
        (BOM_OUT_NOTE_COL, "비고"),
    ]:
        style_header(ws.cell(row=HEADER_TOP, column=col, value=lbl))
        ws.merge_cells(start_row=HEADER_TOP, start_column=col,
                       end_row=HEADER_SUB, end_column=col)

    result: dict[str, dict[int, int]] = {}
    r = HEADER_SUB + 1
    seq = 1
    sum_first_letter = chr(64 + BOM_OUT_MONTH_FIRST_COL)
    sum_last_letter = chr(64 + BOM_OUT_MONTH_LAST_COL)
    for p in products:
        bom = boms_by_product[p.code]
        result[p.code] = {}
        ws.cell(row=r, column=1, value=f"▼ {p.code}")
        ws.cell(row=r, column=1).font = SUBTITLE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        ws.cell(row=r, column=1).fill = LIGHT_GRAY_FILL
        r += 1

        for full_idx, it in enumerate(bom):
            if it.direction != "output":
                continue
            # 폐기물은 별도 시트로 라우팅됨 — BOM 출력물에는 제품만 남김
            if _bom_item_sheet(it) != SHEET_BOM_OUTPUT:
                continue
            ws.cell(row=r, column=1, value=seq); style_body(ws.cell(row=r, column=1), numeric=True, fmt="0")
            ws.cell(row=r, column=2, value=p.code); style_body(ws.cell(row=r, column=2))
            ws.cell(row=r, column=3, value=it.category); style_body(ws.cell(row=r, column=3))
            ws.cell(row=r, column=4, value=it.name); style_body(ws.cell(row=r, column=4))
            ws.cell(row=r, column=5, value=it.applied_unit); style_body(ws.cell(row=r, column=5))
            monthly = it.collected_monthly or [it.applied_qty / 12.0] * 12
            for i, mv in enumerate(monthly):
                c = ws.cell(row=r, column=BOM_OUT_MONTH_FIRST_COL + i, value=mv)
                style_body(c, numeric=True)
            ws.cell(row=r, column=BOM_OUT_SUM_COL,
                    value=f"=SUM({sum_first_letter}{r}:{sum_last_letter}{r})")
            style_body(ws.cell(row=r, column=BOM_OUT_SUM_COL), numeric=True)
            ws.cell(row=r, column=BOM_OUT_SUM_COL).font = HEADER_FONT
            dqi = "M" if (it.dqr_ter <= 2 and it.dqr_ger <= 2) else "C"
            ws.cell(row=r, column=BOM_OUT_DQI_COL, value=dqi)
            style_body(ws.cell(row=r, column=BOM_OUT_DQI_COL))
            ws.cell(row=r, column=BOM_OUT_NOTE_COL, value=it.note or "")
            style_body(ws.cell(row=r, column=BOM_OUT_NOTE_COL))
            if it.category == "제품":
                for col in range(1, last_col + 1):
                    ws.cell(row=r, column=col).fill = LIGHT_GRAY_FILL
            result[p.code][full_idx] = r
            r += 1
            seq += 1

    return result


# ===========================================================================
# 시트 (NEW) — 전기 사용량 (KS I ISO 14067 §6.4.9.4 — 전력 처리 의무)
# ===========================================================================
# 컬럼:
#   A 순번 / B 적용 제품 / C 사용처 / D 전원 구분 (외부그리드/자체발전/직접연결/REC)
#   E 단위 / F~Q 12개월 / R 합계 SUM / S 공급자 / T DQI / U 비고
# ===========================================================================

ELEC_HEADER_ROW = 2
ELEC_DATA_START_ROW = 4   # HEADER_TOP=2, HEADER_SUB=3, 데이터 시작 4
ELEC_MONTH_FIRST_COL = 6  # F
ELEC_MONTH_LAST_COL = 17  # Q
ELEC_SUM_COL = 18         # R
ELEC_SUPPLIER_COL = 19    # S
ELEC_DQI_COL = 20         # T
ELEC_NOTE_COL = 21        # U


def build_electricity(
    ws,
    boms_by_product: dict[str, list[BomItem]],
    products: list[ProductCFP],
) -> dict[str, dict[int, int]]:
    """전기 사용량 전용 시트.
    KS I ISO 14067 §6.4.9.4 (전력 처리) — 그리드/자체/직접연결/REC 4 유형 분류 필수.
    """
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 7
    for c in range(ELEC_MONTH_FIRST_COL, ELEC_MONTH_LAST_COL + 1):
        ws.column_dimensions[chr(64 + c)].width = 9
    ws.column_dimensions[chr(64 + ELEC_SUM_COL)].width = 13
    ws.column_dimensions[chr(64 + ELEC_SUPPLIER_COL)].width = 30
    ws.column_dimensions[chr(64 + ELEC_DQI_COL)].width = 7
    ws.column_dimensions[chr(64 + ELEC_NOTE_COL)].width = 32

    last_col = ELEC_NOTE_COL
    ws.cell(row=1, column=1,
            value="6. 전기 사용량 — 월별 12 컬럼 (KS I ISO 14067 §6.4.9.4)").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    HEADER_TOP = ELEC_HEADER_ROW
    HEADER_SUB = HEADER_TOP + 1
    fixed_left = [
        (1, "순번"), (2, "적용 제품"), (3, "사용처"),
        (4, "전원 구분"), (5, "단위"),
    ]
    for col, lbl in fixed_left:
        style_header(ws.cell(row=HEADER_TOP, column=col, value=lbl))
        ws.merge_cells(start_row=HEADER_TOP, start_column=col,
                       end_row=HEADER_SUB, end_column=col)
    style_header(ws.cell(row=HEADER_TOP, column=ELEC_MONTH_FIRST_COL,
                         value="월별 사용량 (kWh)"))
    ws.merge_cells(start_row=HEADER_TOP, start_column=ELEC_MONTH_FIRST_COL,
                   end_row=HEADER_TOP, end_column=ELEC_MONTH_LAST_COL)
    for i, mlabel in enumerate(KOREAN_MONTHS):
        style_header(ws.cell(row=HEADER_SUB,
                             column=ELEC_MONTH_FIRST_COL + i, value=mlabel))
    for col, lbl in [
        (ELEC_SUM_COL, "합계 (=SUM)"),
        (ELEC_SUPPLIER_COL, "공급자 / EF 출처"),
        (ELEC_DQI_COL, "DQI"),
        (ELEC_NOTE_COL, "비고"),
    ]:
        style_header(ws.cell(row=HEADER_TOP, column=col, value=lbl))
        ws.merge_cells(start_row=HEADER_TOP, start_column=col,
                       end_row=HEADER_SUB, end_column=col)

    result: dict[str, dict[int, int]] = {}
    r = HEADER_SUB + 1
    seq = 1
    sum_first = chr(64 + ELEC_MONTH_FIRST_COL)
    sum_last = chr(64 + ELEC_MONTH_LAST_COL)
    for p in products:
        bom = boms_by_product[p.code]
        result[p.code] = {}
        ws.cell(row=r, column=1, value=f"▼ {p.code}")
        ws.cell(row=r, column=1).font = SUBTITLE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        ws.cell(row=r, column=1).fill = LIGHT_GRAY_FILL
        r += 1

        for full_idx, it in enumerate(bom):
            if _bom_item_sheet(it) != SHEET_ELECTRICITY:
                continue
            ws.cell(row=r, column=1, value=seq); style_body(ws.cell(row=r, column=1), numeric=True, fmt="0")
            ws.cell(row=r, column=2, value=p.code); style_body(ws.cell(row=r, column=2))
            ws.cell(row=r, column=3, value=it.name); style_body(ws.cell(row=r, column=3))
            ws.cell(row=r, column=4, value=it.power_source_type or "외부그리드")
            style_body(ws.cell(row=r, column=4))
            ws.cell(row=r, column=5, value=it.collected_unit); style_body(ws.cell(row=r, column=5))
            monthly = it.collected_monthly or [it.collected_qty / 12.0] * 12
            for i, mv in enumerate(monthly):
                style_body(ws.cell(row=r, column=ELEC_MONTH_FIRST_COL + i, value=mv), numeric=True)
            ws.cell(row=r, column=ELEC_SUM_COL,
                    value=f"=SUM({sum_first}{r}:{sum_last}{r})")
            style_body(ws.cell(row=r, column=ELEC_SUM_COL), numeric=True)
            ws.cell(row=r, column=ELEC_SUM_COL).font = HEADER_FONT
            ws.cell(row=r, column=ELEC_SUPPLIER_COL, value=it.power_supplier or "")
            style_body(ws.cell(row=r, column=ELEC_SUPPLIER_COL))
            dqi = "M" if (it.dqr_ter <= 2 and it.dqr_ger <= 2) else "C"
            ws.cell(row=r, column=ELEC_DQI_COL, value=dqi); style_body(ws.cell(row=r, column=ELEC_DQI_COL))
            ws.cell(row=r, column=ELEC_NOTE_COL, value=it.note or "")
            style_body(ws.cell(row=r, column=ELEC_NOTE_COL))
            result[p.code][full_idx] = r
            r += 1
            seq += 1

    # 검증 노트 — KS I ISO 14067 §6.4.9.4 4 유형 분류
    r += 1
    note_lines = [
        "※ KS I ISO 14067 §6.4.9.4 — 전력 처리 4 유형 분류 의무",
        "  ① 외부그리드: 한국 평균(KECO) EF 적용 / ② 자체발전: 자체 LCI 데이터 수집",
        "  ③ 직접연결: 공급자 특정 EF / ④ REC인증재생: 5.12 중복배제 + 인증서 추적",
    ]
    for line in note_lines:
        c = ws.cell(row=r, column=1, value=line)
        c.font = NOTE_FONT; c.alignment = LEFT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        r += 1

    return result


# ===========================================================================
# 시트 (NEW) — 폐기물 처리 실적
# ===========================================================================
# 컬럼:
#   A 순번 / B 적용 제품 / C 발생공정 / D 폐기물 분류 / E 처리방법
#   F 명칭 / G 단위 / H~S 12개월 / T 합계 SUM
#   U 처리업체 / V 처리장 운송거리 / W DQI / X 비고
# ===========================================================================

WASTE_HEADER_ROW = 2
WASTE_DATA_START_ROW = 4
WASTE_MONTH_FIRST_COL = 8   # H
WASTE_MONTH_LAST_COL = 19   # S
WASTE_SUM_COL = 20          # T
WASTE_FACILITY_COL = 21     # U
WASTE_DIST_COL = 22         # V
WASTE_DQI_COL = 23          # W
WASTE_NOTE_COL = 24         # X


def build_waste(
    ws,
    boms_by_product: dict[str, list[BomItem]],
    products: list[ProductCFP],
) -> dict[str, dict[int, int]]:
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 7
    for c in range(WASTE_MONTH_FIRST_COL, WASTE_MONTH_LAST_COL + 1):
        ws.column_dimensions[chr(64 + c)].width = 9
    ws.column_dimensions[chr(64 + WASTE_SUM_COL)].width = 12
    ws.column_dimensions[chr(64 + WASTE_FACILITY_COL)].width = 24
    ws.column_dimensions[chr(64 + WASTE_DIST_COL)].width = 11
    ws.column_dimensions[chr(64 + WASTE_DQI_COL)].width = 7
    ws.column_dimensions[chr(64 + WASTE_NOTE_COL)].width = 30

    last_col = WASTE_NOTE_COL
    ws.cell(row=1, column=1,
            value="7. 폐기물 처리 실적 — 월별 12 컬럼 + 처리방법/업체/운송"
            ).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    HEADER_TOP = WASTE_HEADER_ROW
    HEADER_SUB = HEADER_TOP + 1
    fixed_left = [
        (1, "순번"), (2, "적용 제품"), (3, "발생공정"),
        (4, "분류"), (5, "처리방법"), (6, "명칭"), (7, "단위"),
    ]
    for col, lbl in fixed_left:
        style_header(ws.cell(row=HEADER_TOP, column=col, value=lbl))
        ws.merge_cells(start_row=HEADER_TOP, start_column=col,
                       end_row=HEADER_SUB, end_column=col)
    style_header(ws.cell(row=HEADER_TOP, column=WASTE_MONTH_FIRST_COL,
                         value="월별 발생량"))
    ws.merge_cells(start_row=HEADER_TOP, start_column=WASTE_MONTH_FIRST_COL,
                   end_row=HEADER_TOP, end_column=WASTE_MONTH_LAST_COL)
    for i, mlabel in enumerate(KOREAN_MONTHS):
        style_header(ws.cell(row=HEADER_SUB,
                             column=WASTE_MONTH_FIRST_COL + i, value=mlabel))
    for col, lbl in [
        (WASTE_SUM_COL, "합계 (=SUM)"),
        (WASTE_FACILITY_COL, "처리업체"),
        (WASTE_DIST_COL, "운송거리(km)"),
        (WASTE_DQI_COL, "DQI"),
        (WASTE_NOTE_COL, "비고"),
    ]:
        style_header(ws.cell(row=HEADER_TOP, column=col, value=lbl))
        ws.merge_cells(start_row=HEADER_TOP, start_column=col,
                       end_row=HEADER_SUB, end_column=col)

    result: dict[str, dict[int, int]] = {}
    r = HEADER_SUB + 1
    seq = 1
    sum_first = chr(64 + WASTE_MONTH_FIRST_COL)
    sum_last = chr(64 + WASTE_MONTH_LAST_COL)
    for p in products:
        bom = boms_by_product[p.code]
        result[p.code] = {}
        ws.cell(row=r, column=1, value=f"▼ {p.code}")
        ws.cell(row=r, column=1).font = SUBTITLE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        ws.cell(row=r, column=1).fill = LIGHT_GRAY_FILL
        r += 1

        for full_idx, it in enumerate(bom):
            if _bom_item_sheet(it) != SHEET_WASTE:
                continue
            ws.cell(row=r, column=1, value=seq); style_body(ws.cell(row=r, column=1), numeric=True, fmt="0")
            ws.cell(row=r, column=2, value=p.code); style_body(ws.cell(row=r, column=2))
            ws.cell(row=r, column=3, value="제품 생산 공정"); style_body(ws.cell(row=r, column=3))
            ws.cell(row=r, column=4, value=it.category); style_body(ws.cell(row=r, column=4))
            ws.cell(row=r, column=5, value=it.treatment_method or ""); style_body(ws.cell(row=r, column=5))
            ws.cell(row=r, column=6, value=it.name); style_body(ws.cell(row=r, column=6))
            ws.cell(row=r, column=7, value=it.applied_unit); style_body(ws.cell(row=r, column=7))
            monthly = it.collected_monthly or [it.applied_qty / 12.0] * 12
            for i, mv in enumerate(monthly):
                style_body(ws.cell(row=r, column=WASTE_MONTH_FIRST_COL + i, value=mv), numeric=True)
            ws.cell(row=r, column=WASTE_SUM_COL,
                    value=f"=SUM({sum_first}{r}:{sum_last}{r})")
            style_body(ws.cell(row=r, column=WASTE_SUM_COL), numeric=True)
            ws.cell(row=r, column=WASTE_SUM_COL).font = HEADER_FONT
            ws.cell(row=r, column=WASTE_FACILITY_COL, value=it.treatment_facility or "")
            style_body(ws.cell(row=r, column=WASTE_FACILITY_COL))
            if it.treatment_distance_km:
                ws.cell(row=r, column=WASTE_DIST_COL, value=it.treatment_distance_km)
                style_body(ws.cell(row=r, column=WASTE_DIST_COL), numeric=True)
            else:
                ws.cell(row=r, column=WASTE_DIST_COL, value="")
                style_body(ws.cell(row=r, column=WASTE_DIST_COL))
            dqi = "M" if (it.dqr_ter <= 2 and it.dqr_ger <= 2) else "C"
            ws.cell(row=r, column=WASTE_DQI_COL, value=dqi); style_body(ws.cell(row=r, column=WASTE_DQI_COL))
            ws.cell(row=r, column=WASTE_NOTE_COL, value=it.note or ""); style_body(ws.cell(row=r, column=WASTE_NOTE_COL))
            result[p.code][full_idx] = r
            r += 1
            seq += 1

    return result


# ===========================================================================
# 시트 5 — 제품별 CFP (sheet5 'E2B-03E' 레이아웃 단순화 차용)
#   - input/output 테이블, EF×환산량 수식, DQR 5축 (KS I ISO 14067 §6.3.5),
#     cut-off 누적 기여도 표 (§6.3.4.3 제외 기준)
#   - "살아있는 수식" 부분 적용: S=L×N, N=G/$G$FU, T=S/$S$total, X=$T×U
# ===========================================================================

PRODUCT_CFP_HEADERS = [
    # (col_idx, top_label, sub_label, width)
    (1, "구분", "", 8),       # A
    (2, "분류", "", 12),      # B
    (3, "명칭", "", 36),      # C
    (4, "수집 단위", "데이터 수집", 8),    # D
    (5, "수집 수량", "데이터 수집", 12),   # E
    (6, "적용 단위", "데이터 적용", 8),    # F
    (7, "적용 수량", "데이터 적용", 12),   # G
    (8, "cut-off rule", "", 32),           # H
    (9, "activity name", "", 36),          # I
    (10, "flow name", "", 22),             # J
    (11, "Location", "", 10),              # K
    (12, "EF (kgCO₂e/unit)", "", 12),      # L
    (13, "FU 단위", "제품", 8),            # M
    (14, "FU 환산량", "제품", 12),         # N
    (15, "운송수단", "운송", 12),          # O
    (16, "운송거리(km)", "운송", 10),      # P
    (17, "ton-km", "운송", 10),            # Q
    (18, "비고", "", 28),                  # R
    (19, "온실가스 배출량 (kgCO₂e)", "", 14),  # S
    (20, "기여도", "", 9),                 # T
    (21, "TeR", "DQR", 5),                # U
    (22, "GeR", "DQR", 5),                # V
    (23, "TiR", "DQR", 5),                # W
    (24, "TeR×T", "가중평균 DQR", 7),      # X
    (25, "GeR×T", "가중평균 DQR", 7),      # Y
    (26, "TiR×T", "가중평균 DQR", 7),      # Z
]


SHEET_PRODUCTION = "제품 생산량"
SHEET_BOM_INPUT = "BOM 입력물"
SHEET_BOM_OUTPUT = "BOM 출력물"
SHEET_ELECTRICITY = "전기 사용량"
SHEET_WASTE = "폐기물 처리 실적"
SHEET_EF_DB = "사용한 2차 데이터 목록"
SHEET_LCIA = "LCIA"


# BomItem → 어느 시트에 들어가는지 분류
#   '전기 사용량'   : input + 카테고리 '에너지' + 명칭에 '전력' 포함
#   '폐기물 처리 실적': output + 카테고리 in {매립, 지정폐기물, 폐수, 소각, 재활용}
#   'BOM 입력물'   : 기타 input
#   'BOM 출력물'   : 기타 output (실질적으로 제품 행만)
WASTE_CATEGORIES = {"매립", "지정폐기물", "폐수", "소각", "재활용", "위탁처리"}


def _bom_item_sheet(it: BomItem) -> str:
    if it.direction == "input":
        if it.category == "에너지" and "전력" in it.name:
            return SHEET_ELECTRICITY
        return SHEET_BOM_INPUT
    if it.category in WASTE_CATEGORIES:
        return SHEET_WASTE
    return SHEET_BOM_OUTPUT

# EF DB 시트의 헤더 위치 (build_secondary_data 참조)
EF_DB_HEADER_ROW = 7
EF_DB_DATA_START = 8


def _ef_db_row(ef_seq: int) -> int:
    """ef_seq (1-base) 를 EF DB 시트의 절대 행 번호로 변환."""
    return EF_DB_DATA_START + (ef_seq - 1)


def build_product_cfp(
    ws,
    *,
    product: ProductCFP,
    bom: list[BomItem],
    bom_input_rows: dict[int, int],
    bom_output_rows: dict[int, int],
    electricity_rows: dict[int, int] | None = None,
    waste_rows: dict[int, int] | None = None,
    fu_anchor_row: int = PRODUCTION_DATA_START,
) -> int:
    """제품별 CFP 시트. 토리컴(NiSO4) 1제품용 단순화 + Tier 1·2 cross-ref 버전.

    원본 sheet5 (135행 × 34열, 691 수식) 의 핵심 구조 차용:
      - 모든 활동량은 BOM 입력물/출력물 시트 참조 (`='BOM 입력물'!H{n}`)
      - 모든 EF/메타는 EF DB 시트 참조 (`='사용한 2차 데이터 목록'!J{n}`)
      - FU 분모는 제품 생산량 시트 참조 (`='제품 생산량'!E3`)
      - 합계/기여도/DQR 가중평균은 자체 시트 내 수식

    반환: total_row (LCIA 시트가 합계 셀을 참조할 수 있도록)
    """
    # 컬럼 너비
    for idx, _top, _sub, w in PRODUCT_CFP_HEADERS:
        ws.column_dimensions[chr(64 + idx)].width = w

    # 제목
    ws.cell(row=1, column=1, value=f"{product.display_name} 탄소발자국 산정 데이터").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=26)

    # 헤더 (2~3행, 머지)
    HEADER_TOP, HEADER_SUB = 2, 3
    # 그룹 머지 정보 — (start_col, end_col, label)
    grouped = {
        "데이터 수집": (4, 5),
        "데이터 적용": (6, 7),
        "제품": (13, 14),
        "운송": (15, 17),
        "DQR": (21, 23),
        "가중평균 DQR": (24, 26),
    }
    placed_groups = set()
    for idx, top, sub, _w in PRODUCT_CFP_HEADERS:
        if sub and sub in grouped and sub not in placed_groups:
            sc, ec = grouped[sub]
            ws.cell(row=HEADER_TOP, column=sc, value=sub)
            style_header(ws.cell(row=HEADER_TOP, column=sc))
            if sc != ec:
                ws.merge_cells(start_row=HEADER_TOP, start_column=sc,
                               end_row=HEADER_TOP, end_column=ec)
            placed_groups.add(sub)
            # sub 라벨이 있는 컬럼은 sub 행에 세부 라벨
            ws.cell(row=HEADER_SUB, column=idx, value=top)
            style_header(ws.cell(row=HEADER_SUB, column=idx))
        elif sub and sub in placed_groups:
            # 같은 그룹의 후속 컬럼 — sub 행에만 세부
            ws.cell(row=HEADER_SUB, column=idx, value=top)
            style_header(ws.cell(row=HEADER_SUB, column=idx))
        else:
            # 단일 컬럼 — 2~3행 머지
            ws.cell(row=HEADER_TOP, column=idx, value=top)
            style_header(ws.cell(row=HEADER_TOP, column=idx))
            ws.merge_cells(start_row=HEADER_TOP, start_column=idx,
                           end_row=HEADER_SUB, end_column=idx)

    # 데이터 행 시작
    DATA_START = 4
    r = DATA_START

    # FU anchor row 위치를 미리 결정 (output 제품 행 — N 열 분모 $G$X 참조용)
    fu_row = None
    for i, item in enumerate(bom):
        if item.direction == "output" and item.category == "제품":
            fu_row = DATA_START + i
            break
    if fu_row is None:
        raise ValueError("BOM에 output 제품 행이 없습니다 (FU anchor 필요)")

    # 합계행 위치를 미리 계산 (S/T/X/Y/Z 수식이 row 번호를 참조해야 함)
    total_row = DATA_START + len(bom)

    # 제품 생산량 Q열 (합계 SUM 셀) 이 FU 분모. monthly 추가로 컬럼 letter 변경됨.
    fu_anchor_ref = f"'{SHEET_PRODUCTION}'!{PROD_FU_ANCHOR_LETTER}{fu_anchor_row}"

    electricity_rows = electricity_rows or {}
    waste_rows = waste_rows or {}

    for idx, item in enumerate(bom):
        # 항목 분류에 따라 4개 시트 중 하나를 cross-ref:
        #   전기 사용량      → R (합계 SUM), U(=비고)
        #   폐기물 처리 실적 → T (합계 SUM), X (비고)
        #   BOM 입력물      → R (합계), U (적용), X (비고)
        #   BOM 출력물      → R (합계), R (적용), T (비고)
        target_sheet = _bom_item_sheet(item)
        if target_sheet == SHEET_ELECTRICITY:
            bom_row = electricity_rows[idx]
            bom_sheet = SHEET_ELECTRICITY
            collected_col = chr(64 + ELEC_SUM_COL)       # R
            applied_col = chr(64 + ELEC_SUM_COL)         # R (전력은 농도 미적용)
        elif target_sheet == SHEET_WASTE:
            bom_row = waste_rows[idx]
            bom_sheet = SHEET_WASTE
            collected_col = chr(64 + WASTE_SUM_COL)      # T
            applied_col = chr(64 + WASTE_SUM_COL)        # T
        elif item.direction == "input":
            bom_row = bom_input_rows[idx]
            bom_sheet = SHEET_BOM_INPUT
            collected_col = chr(64 + BOM_IN_SUM_COL)     # R
            applied_col = chr(64 + BOM_IN_QTY_COL)       # U
        else:
            bom_row = bom_output_rows[idx]
            bom_sheet = SHEET_BOM_OUTPUT
            collected_col = chr(64 + BOM_OUT_SUM_COL)    # R
            applied_col = chr(64 + BOM_OUT_SUM_COL)      # R

        # A 구분 / B 분류 / C 명칭
        ws.cell(row=r, column=1, value=item.direction)
        ws.cell(row=r, column=2, value=item.category)
        ws.cell(row=r, column=3, value=item.name)
        # D-E 데이터 수집 (단위 + 수식)
        ws.cell(row=r, column=4, value=item.collected_unit)
        ws.cell(row=r, column=5, value=f"='{bom_sheet}'!{collected_col}{bom_row}")
        # F-G 데이터 적용 (단위 + 수식)
        ws.cell(row=r, column=6, value=item.applied_unit)
        ws.cell(row=r, column=7, value=f"='{bom_sheet}'!{applied_col}{bom_row}")
        # H cut-off
        ws.cell(row=r, column=8, value=item.cut_off if item.cut_off else "")

        # I/J/K/L — EF DB 참조 (cut-off / FU 행은 비움)
        if item.ef_seq > 0:
            ef_row = _ef_db_row(item.ef_seq)
            ws.cell(row=r, column=9, value=f"='{SHEET_EF_DB}'!E{ef_row}")   # activity name
            ws.cell(row=r, column=10, value=f"='{SHEET_EF_DB}'!G{ef_row}")  # flow name
            ws.cell(row=r, column=11, value=f"='{SHEET_EF_DB}'!F{ef_row}")  # location
            ws.cell(row=r, column=12, value=f"='{SHEET_EF_DB}'!J{ef_row}")  # EF
        else:
            ws.cell(row=r, column=9, value="—")
            ws.cell(row=r, column=10, value="—")
            ws.cell(row=r, column=11, value=item.location or "—")
            ws.cell(row=r, column=12, value=0)

        # M FU 단위
        ws.cell(row=r, column=13, value="kg")
        # N FU 환산량
        if r == fu_row:
            ws.cell(row=r, column=14, value=1)
        else:
            ws.cell(row=r, column=14, value=f"=G{r}/{fu_anchor_ref}")
        # O-Q 운송
        ws.cell(row=r, column=15, value=item.transport_mode)
        if item.transport_distance_km:
            # 운송거리: 입력 BOM V열 (BOM_IN_DIST_COL)
            dist_letter = chr(64 + BOM_IN_DIST_COL)
            ws.cell(row=r, column=16, value=f"='{SHEET_BOM_INPUT}'!{dist_letter}{bom_row}")
            ws.cell(row=r, column=17, value=f"=N{r}*P{r}/1000")
        else:
            ws.cell(row=r, column=16, value=None)
            ws.cell(row=r, column=17, value=None)
        # R 비고 — 시트별 비고 컬럼 매핑
        if target_sheet == SHEET_ELECTRICITY:
            note_letter = chr(64 + ELEC_NOTE_COL)    # U
        elif target_sheet == SHEET_WASTE:
            note_letter = chr(64 + WASTE_NOTE_COL)   # X
        elif item.direction == "input":
            note_letter = chr(64 + BOM_IN_NOTE_COL)  # X
        else:
            note_letter = chr(64 + BOM_OUT_NOTE_COL) # T
        ws.cell(row=r, column=18, value=f"='{bom_sheet}'!{note_letter}{bom_row}")
        # S 배출량 (L × N), 단 cut-off / FU 는 0
        if item.direction == "output" and item.category == "제품":
            ws.cell(row=r, column=19, value=0)
        elif item.cut_off:
            ws.cell(row=r, column=19, value=0)
        else:
            ws.cell(row=r, column=19, value=f"=L{r}*N{r}")
        # T 기여도
        if item.direction == "output" and item.category == "제품":
            ws.cell(row=r, column=20, value="—")
        else:
            ws.cell(row=r, column=20, value=f"=IFERROR(S{r}/$S${total_row},0)")
        # DQR
        ws.cell(row=r, column=21, value=item.dqr_ter)
        ws.cell(row=r, column=22, value=item.dqr_ger)
        ws.cell(row=r, column=23, value=item.dqr_tir)
        if item.direction == "output" and item.category == "제품":
            for col in (24, 25, 26):
                ws.cell(row=r, column=col, value="—")
        else:
            ws.cell(row=r, column=24, value=f"=IFERROR($T{r}*U{r},0)")
            ws.cell(row=r, column=25, value=f"=IFERROR($T{r}*V{r},0)")
            ws.cell(row=r, column=26, value=f"=IFERROR($T{r}*W{r},0)")

        # 셀 스타일링
        for col in range(1, 27):
            c = ws.cell(row=r, column=col)
            numeric = col in (5, 7, 12, 14, 16, 17, 19, 20, 21, 22, 23, 24, 25, 26)
            fmt = "#,##0.00###"
            if col == 20:  # 기여도 %
                fmt = "0.00%"
            elif col in (21, 22, 23):  # DQR 정수
                fmt = "0"
            elif col == 12:  # EF
                fmt = "#,##0.0000"
            style_body(c, numeric=numeric, fmt=fmt)
            # cut-off 행 강조
            if item.cut_off and col in range(1, 19):
                c.fill = WARN_FILL
            # 제품(FU anchor) 행 연한 강조
            if item.direction == "output" and item.category == "제품":
                c.fill = LIGHT_GRAY_FILL
        r += 1

    # 합계행
    assert r == total_row
    ws.cell(row=r, column=1, value="합계").font = HEADER_FONT
    style_header(ws.cell(row=r, column=1))
    ws.cell(row=r, column=2, value=""); style_header(ws.cell(row=r, column=2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=18)
    # S 합계 — input 행만 합산 (output 폐기물은 제외해야 하나, sheet5 패턴 따라 input + waste output 모두 합산)
    # sheet5의 S53 = SUM(S4:S42, S47:S49, S52). 토리컴 단순화: 모든 input + 폐기물 output 합산 (제품 행은 0이므로 포함 무방)
    ws.cell(row=r, column=19, value=f"=SUM(S{DATA_START}:S{total_row - 1})")
    style_body(ws.cell(row=r, column=19), numeric=True)
    ws.cell(row=r, column=19).font = HEADER_FONT
    ws.cell(row=r, column=20, value=f"=IFERROR(S{r}/$S${r},0)")
    style_body(ws.cell(row=r, column=20), numeric=True, fmt="0.00%")
    # DQR 가중합
    for col_letter, col_idx in (("X", 24), ("Y", 25), ("Z", 26)):
        ws.cell(row=r, column=col_idx, value=f"=SUM({col_letter}{DATA_START}:{col_letter}{total_row - 1})")
        style_body(ws.cell(row=r, column=col_idx), numeric=True)
    # AA 평균 DQR
    ws.cell(row=r, column=27, value=f"=AVERAGE(X{r}:Z{r})")
    style_body(ws.cell(row=r, column=27), numeric=True)
    ws.column_dimensions["AA"].width = 11
    ws.cell(row=r, column=21).value = "전체 DQR (가중평균)"
    style_header(ws.cell(row=r, column=21))
    ws.merge_cells(start_row=r, start_column=21, end_row=r, end_column=23)

    # 검증 요약 블록
    r += 2
    ws.cell(row=r, column=1, value="검증 요약 (KS I ISO 14067 §6.3.4.3 / §5.7)").font = SUBTITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    summary = [
        ("제품 (Functional Unit)", f"{product.fu_label} ({product.functional_unit})"),
        ("총 CFP", f"=S{total_row}"),
        ("CFP / FU (kgCO₂e/ton)", f"=S{total_row}*1000/G{fu_row}"),  # 1ton = 1000kg
        ("입력 항목 수 (cut-off 포함)", str(sum(1 for it in bom if it.direction == "input"))),
        ("Cut-off 적용 항목 수", str(sum(1 for it in bom if it.cut_off))),
        ("출력 항목 수 (제품 + 폐기물)", str(sum(1 for it in bom if it.direction == "output"))),
    ]
    for label, value in summary:
        ws.cell(row=r, column=1, value=label); style_header(ws.cell(row=r, column=1))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(row=r, column=5, value=value)
        if isinstance(value, str) and value.startswith("="):
            style_body(ws.cell(row=r, column=5), numeric=True)
        else:
            style_body(ws.cell(row=r, column=5))
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10)
        r += 1

    # Cut-off 누적 기여도 표 (KS I ISO 14067 §6.3.4.3 제외 기준)
    r += 1
    ws.cell(row=r, column=1, value="Cut-off 누적 질량 기여도 — §6.3.4.3 제외 기준").font = SUBTITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    for col, label in enumerate(["순번", "분류", "명칭", "단위", "투입량", "비율(%)", "누적 질량 기여도(%)", "비고"], start=1):
        style_header(ws.cell(row=r, column=col, value=label))
    r += 1
    cutoff_start = r
    # input 만 추출하여 mass-based 정렬
    inputs_with_mass = [it for it in bom if it.direction == "input"]
    # 질량 기준 정렬 (kg 단위만 mass-based, m3/kWh/Nm3/ton-km 는 mass-based 비교에서 제외)
    mass_items = [it for it in inputs_with_mass if it.applied_unit == "kg"]
    mass_items.sort(key=lambda x: x.applied_qty, reverse=True)
    # 비-mass 항목은 따로 표기 (전력/물/LNG/스팀/운송)
    non_mass_items = [it for it in inputs_with_mass if it.applied_unit != "kg"]

    seq = 1
    for it in mass_items:
        ws.cell(row=r, column=1, value=seq); style_body(ws.cell(row=r, column=1), numeric=True, fmt="0")
        ws.cell(row=r, column=2, value=it.category); style_body(ws.cell(row=r, column=2))
        ws.cell(row=r, column=3, value=it.name); style_body(ws.cell(row=r, column=3))
        ws.cell(row=r, column=4, value=it.applied_unit); style_body(ws.cell(row=r, column=4), numeric=True)
        ws.cell(row=r, column=5, value=it.applied_qty); style_body(ws.cell(row=r, column=5), numeric=True)
        # 비율 = E_r / SUM(E_cutoff_start:E_lastmass)
        ws.cell(row=r, column=6, value=f"=E{r}/SUM($E${cutoff_start}:$E${cutoff_start + len(mass_items) - 1})")
        style_body(ws.cell(row=r, column=6), numeric=True, fmt="0.00%")
        # 누적 = 위 비율 + 직전 누적 (첫 행은 자기 자신)
        if r == cutoff_start:
            ws.cell(row=r, column=7, value=f"=F{r}")
        else:
            ws.cell(row=r, column=7, value=f"=G{r - 1}+F{r}")
        style_body(ws.cell(row=r, column=7), numeric=True, fmt="0.00%")
        ws.cell(row=r, column=8, value=it.cut_off if it.cut_off else "포함"); style_body(ws.cell(row=r, column=8))
        if it.cut_off:
            for col in range(1, 9):
                ws.cell(row=r, column=col).fill = WARN_FILL
        r += 1
        seq += 1
    # 비-mass 항목 (참고)
    if non_mass_items:
        ws.cell(row=r, column=1, value="—"); style_body(ws.cell(row=r, column=1))
        ws.cell(row=r, column=2, value="비질량 입력 (참고)"); style_body(ws.cell(row=r, column=2))
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        r += 1
        for it in non_mass_items:
            ws.cell(row=r, column=1, value=seq); style_body(ws.cell(row=r, column=1), numeric=True, fmt="0")
            ws.cell(row=r, column=2, value=it.category); style_body(ws.cell(row=r, column=2))
            ws.cell(row=r, column=3, value=it.name); style_body(ws.cell(row=r, column=3))
            ws.cell(row=r, column=4, value=it.applied_unit); style_body(ws.cell(row=r, column=4), numeric=True)
            ws.cell(row=r, column=5, value=it.applied_qty); style_body(ws.cell(row=r, column=5), numeric=True)
            ws.cell(row=r, column=6, value="—"); style_body(ws.cell(row=r, column=6))
            ws.cell(row=r, column=7, value="—"); style_body(ws.cell(row=r, column=7))
            ws.cell(row=r, column=8, value="질량 기준 비교 대상 외 — 별도 평가"); style_body(ws.cell(row=r, column=8))
            r += 1
            seq += 1

    # ---------------------------------------------------------------
    # K — 차트 2종 (살아있는 셀 참조)
    #   1) 질량 기여도 Bar (cut-off 표 — §6.3.4.3 95% 검증 시각화)
    #   2) 항목별 DQR Bar (TeR/GeR/TiR — §6.3.5 데이터 품질 시각화)
    # ---------------------------------------------------------------
    chart_row = r + 1
    if mass_items:
        mass_chart = BarChart()
        mass_chart.type = "bar"  # 가로 막대
        mass_chart.style = 11
        mass_chart.title = "원료 질량 기여도 (Cut-off 95% 검증)"
        mass_chart.x_axis.title = "투입량 (kg)"
        mass_first = cutoff_start
        mass_last = cutoff_start + len(mass_items) - 1
        # 헤더 라벨은 cutoff_start - 1 행, '명칭' = 3컬럼, '투입량' = 5컬럼
        cats = Reference(ws, min_col=3, min_row=mass_first, max_row=mass_last)
        vals = Reference(ws, min_col=5, min_row=mass_first - 1, max_row=mass_last)
        mass_chart.add_data(vals, titles_from_data=True)
        mass_chart.set_categories(cats)
        mass_chart.dataLabels = DataLabelList(showVal=True)
        mass_chart.height = 8
        mass_chart.width = 16
        ws.add_chart(mass_chart, f"J{chart_row}")

    # DQR 차트 — 항목별 TeR/GeR/TiR (3 시리즈)
    dqr_chart = BarChart()
    dqr_chart.type = "col"
    dqr_chart.style = 12
    dqr_chart.title = "항목별 DQR 점수 (1=best ~ 5=worst, KS I ISO 14067 §6.3.5)"
    dqr_chart.y_axis.scaling.min = 0
    dqr_chart.y_axis.scaling.max = 5
    # DATA_START = 4, last data row = total_row - 1 = 20
    dqr_first = DATA_START
    dqr_last = total_row - 1
    cats = Reference(ws, min_col=3, min_row=dqr_first, max_row=dqr_last)
    # U(21)~W(23) = TeR/GeR/TiR. 헤더가 row 3 (sub-label 행)
    vals = Reference(ws, min_col=21, max_col=23,
                     min_row=dqr_first - 1, max_row=dqr_last)
    dqr_chart.add_data(vals, titles_from_data=True)
    dqr_chart.set_categories(cats)
    dqr_chart.height = 8
    dqr_chart.width = 18
    # Pareto 차트 아래에 위치
    ws.add_chart(dqr_chart, f"J{chart_row + 17}")

    # 행 높이 자동 조정 (헤더만)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    # 인쇄 — 가로 방향
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    return total_row


# ===========================================================================
# 워크북 빌더
# ===========================================================================

def _build_stage_row_map(bom: list[BomItem], product_data_start: int) -> dict[str, list[int]]:
    """전과정 단계명 → 해당 단계에 속하는 product CFP 시트 행 번호 리스트.

    토리컴 단계 매핑 (cfp-result run05 기준):
      - '원료 채취'  : input.원료물질 + input.유틸리티
      - '제조'      : input.에너지 + output.매립 + output.지정폐기물 + output.폐수
      - '운송'      : input.육상운송
      - '포장'      : input.포장
    """
    mapping = {
        "원료 채취": ["원료물질", "유틸리티"],
        "제조": ["에너지", "매립", "지정폐기물", "폐수"],
        "운송": ["육상운송"],
        "포장": ["포장"],
    }
    result: dict[str, list[int]] = {stage: [] for stage in mapping}
    for idx, item in enumerate(bom):
        row = product_data_start + idx
        for stage, cats in mapping.items():
            if item.category in cats:
                # cut-off 행은 S=0이므로 포함해도 합계 영향 없음
                result[stage].append(row)
                break
    return result


def build_workbook(
    *,
    meta: StudyMeta,
    products: list[ProductCFP],
    sensitivity: list[SensitivityScenario],
    secondary: list[SecondaryDataItem],
    product_boms: dict[str, list[BomItem]],
) -> Workbook:
    wb = Workbook()

    # 시트 1: 표지
    ws_cover = wb.active
    ws_cover.title = "표지"
    build_cover(ws_cover, meta)

    # 시트 2: 제품 생산량 (Tier 1 — FU anchor, 다제품 + 월별 12 컬럼)
    ws_prod = wb.create_sheet(SHEET_PRODUCTION)
    fu_kg_by_product = {p.code: 1000.0 for p in products}
    # 제품별 월별 생산 패턴 (가상) — 각 제품 BOM 의 weights 와 일관된 패턴을 사용
    monthly_weights_by_product = {
        p.code: (_MONTHLY_WEIGHTS_POWDER if "POWDER" in p.code.upper()
                 or "Powder" in p.code else _MONTHLY_WEIGHTS_NISO4)
        for p in products
    }
    fu_anchor_rows = build_production(
        ws_prod, products, fu_kg_by_product,
        monthly_weights_by_product=monthly_weights_by_product,
    )

    # 시트 3-4: BOM 입력물/출력물 (Tier 1, 원료/유틸/연료/스팀/운송/포장 + 제품)
    ws_bom_in = wb.create_sheet(SHEET_BOM_INPUT)
    bom_input_rows_by_product = build_bom_input(ws_bom_in, product_boms, products)
    ws_bom_out = wb.create_sheet(SHEET_BOM_OUTPUT)
    bom_output_rows_by_product = build_bom_output(ws_bom_out, product_boms, products)

    # 시트 5-6: 전기 사용량 + 폐기물 처리 실적 (Tier 1 — 별도 시트, ISO 14067 §6.4.9.4 / §6.3.8 준수)
    ws_elec = wb.create_sheet(SHEET_ELECTRICITY)
    electricity_rows_by_product = build_electricity(ws_elec, product_boms, products)
    ws_waste = wb.create_sheet(SHEET_WASTE)
    waste_rows_by_product = build_waste(ws_waste, product_boms, products)

    # 시트 7: 사용한 2차 데이터 목록 (Tier 2)
    ws_sec = wb.create_sheet(SHEET_EF_DB)
    build_secondary_data(ws_sec, secondary)

    # 시트 8+: 제품별 CFP (Tier 3 — 산출물 본체)
    product_sheet_names: dict[str, str] = {}
    product_total_rows: dict[str, int] = {}
    stage_row_maps: dict[str, dict[str, list[int]]] = {}
    for p in products:
        bom = product_boms.get(p.code)
        if not bom:
            continue
        sheet_name = p.code[:31]
        ws_p = wb.create_sheet(sheet_name)
        total_row = build_product_cfp(
            ws_p,
            product=p,
            bom=bom,
            bom_input_rows=bom_input_rows_by_product[p.code],
            bom_output_rows=bom_output_rows_by_product[p.code],
            electricity_rows=electricity_rows_by_product.get(p.code, {}),
            waste_rows=waste_rows_by_product.get(p.code, {}),
            fu_anchor_row=fu_anchor_rows[p.code],
        )
        product_sheet_names[p.code] = sheet_name
        product_total_rows[p.code] = total_row
        stage_row_maps[p.code] = _build_stage_row_map(bom, product_data_start=4)

    # 시트 7: LCIA (Tier 4 — 요약, product CFP 참조)
    ws_lcia = wb.create_sheet(SHEET_LCIA)
    build_lcia(
        ws_lcia, products,
        product_sheet_names=product_sheet_names,
        product_total_rows=product_total_rows,
        stage_row_maps=stage_row_maps,
    )

    # 시트 8: 민감도 분석 (Tier 4)
    ws_sens = wb.create_sheet("민감도 분석")
    build_sensitivity(ws_sens, product=products[0], scenarios=sensitivity)

    # 인쇄 옵션 일괄 적용
    for ws in wb.worksheets:
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.outlinePr.summaryBelow = False

    return wb


# ===========================================================================
# main
# ===========================================================================

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = REPO_ROOT / "검증사례" / "poc"
OUT_PATH = OUT_DIR / "toricomm_calc_workbook.xlsx"


def _resolve_out_path(base: Path) -> Path:
    """기본 경로에 저장. 파일이 Excel에서 열려 락이 걸렸다면 단일 timestamp suffix 로 fallback.
    숫자 버전(_v2, _v3...) 누적 방식은 폐기 — 디스크 오염 방지.
    """
    if not base.exists():
        return base
    try:
        with open(base, "ab"):
            pass
        return base
    except PermissionError:
        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        cand = base.with_stem(f"{base.stem}_{ts}")
        print(
            f"[WARN] 기본 출력 파일이 Excel에서 열려 있어 락 상태입니다.\n"
            f"       임시 파일에 저장합니다: {cand.name}\n"
            f"       정상화하려면: 1) Excel을 닫고  2) 이 임시 파일을 기본 파일에 덮어쓰기 또는 본 스크립트 재실행."
        )
        try:
            with open(cand, "ab"):
                pass
        except PermissionError:
            raise RuntimeError(
                f"임시 파일도 잠겨 있습니다: {cand.name}\n"
                "Excel 인스턴스가 여러 개 열려 있는지 확인하세요."
            )
        return cand


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(
        meta=TORICOMM_META,
        products=[TORICOMM_NISO4, TORICOMM_NISO4_POWDER_HYPO],
        sensitivity=TORICOMM_SENSITIVITY,
        secondary=TORICOMM_SECONDARY,
        product_boms={
            TORICOMM_NISO4.code: TORICOMM_BOM,
            TORICOMM_NISO4_POWDER_HYPO.code: TORICOMM_BOM_POWDER_HYPO,
        },
    )
    out_path = _resolve_out_path(OUT_PATH)
    wb.save(out_path)
    print(f"[OK] saved: {out_path}")
    print(f"      sheets: {[ws.title for ws in wb.worksheets]}")
    print(f"      product: {TORICOMM_NISO4.display_name}")
    print(f"      total:   {TORICOMM_NISO4.total} {TORICOMM_NISO4.unit} / {TORICOMM_NISO4.functional_unit}")
    print(f"      sensitivity scenarios: {len(TORICOMM_SENSITIVITY)}")
    print(f"      secondary data items:  {len(TORICOMM_SECONDARY)}")
    print(f"      BOM items:             {len(TORICOMM_BOM)}")


if __name__ == "__main__":
    main()
